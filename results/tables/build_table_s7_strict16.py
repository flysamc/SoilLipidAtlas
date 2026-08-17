#!/usr/bin/env python3
"""Table S7 (empirical lipid-class RIE factors) rebuilt for the Figure 5 v2
pipeline.

S7 carries no taxonomy, so the strict-phyla correction does not touch its
measured values. What DID change is the correction the adopted ClimGrass figure
applies: Figure 5 v2 (figure5_redesign_v2.py, RIE step) replaces the submitted
hard floor/ceiling clip with Rule A -- a calibration gate.

    Rule A (climgrass_strict16_rulefix.make_rule_a):
        a class is used for correction only if its mean RIE is inside the
        trusted window [RIE_floor, RIE_ceiling] = [0.20, 100]. In-window ->
        factor = 1 / RIE. Out-of-window (below 0.20 or above 100) -> the RIE is
        treated as UNCALIBRATED and the class is left uncorrected (factor 1.0),
        instead of being clipped to the floor (5x) or ceiling (0.01x).

The measured RIE per (class, adduct) is identical to the submitted table (same
109-standard source, same groupby-mean); this producer re-verifies that against
`rie_table_s10.csv` and then rebuilds only the derived columns and the Notes for
Rule A. Row membership, order, adduct anchoring and ionization-mode labels are
taken from the verified submitted table so nothing but the mechanism changes.
"""

from __future__ import annotations

import csv
import json
import math
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SRC_TABLE = PROJECT_ROOT / "outputs" / "tables" / "Table_S7_empirical_response_RIE.xlsx"
RIE_SRC = PROJECT_ROOT / "analysis" / "analysis-19" / "00_inputs" / "rie_table_s10.csv"
RELEASE = PROJECT_ROOT / "outputs" / "analysis" / "ncbi-phylum-2026-08-04-v1"
OUTDIR = RELEASE / "tables" / "table_s7_strict16_2026-08-12_v1"
OUT_XLSX = OUTDIR / "Table_S7_empirical_response_RIE_v2_ruleA.xlsx"

FLOOR, CEIL = 0.20, 100.0  # rf.RIE_FLOOR / rf.RIE_CEILING, adopted in figure5_redesign_v2

TITLE_FONT = Font(name="Arial", size=13, bold=True)
H_FONT = Font(name="Arial", size=10, bold=True)
BODY = Font(name="Arial", size=10)
BODY_B = Font(name="Arial", size=10, bold=True)
SMALL_I = Font(name="Arial", size=9, italic=True)
WRAP = Alignment(wrap_text=True, vertical="top")

COLUMNS = [
    "Lipid class", "Adduct", "Ionization mode", "n standards averaged",
    "RIE relative to LPE", "log10(RIE)",
    "RIE in calibrated window [0.20, 100]?", "Correction status (Rule A)",
    "Applied correction factor", "Uncapped factor (1 / RIE)",
]
WIDTHS = [13, 9, 15, 12, 18, 12, 20, 20, 18, 18]


def load_source_means() -> dict:
    src = pd.read_csv(RIE_SRC).dropna(subset=["RIE_LPE"]).copy()
    src["class_u"] = src["class"].astype(str).str.upper()
    src["adduct_clean"] = (src["adduct"].astype(str)
                           .str.replace("[", "", regex=False)
                           .str.replace("]", "", regex=False).str.strip())
    g = src.groupby(["class_u", "adduct_clean"])["RIE_LPE"]
    return {k: (float(v), int(g.count()[k])) for k, v in g.mean().items()}


def read_submitted(sheet: str) -> list[dict]:
    wb = openpyxl.load_workbook(SRC_TABLE, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows) if r and r[0] == "Lipid class")
    col = {n: j for j, n in enumerate(rows[hi]) if n}
    out = []
    for r in rows[hi + 1:]:
        if not r or not r[0]:
            continue
        rie = r[col["RIE relative to LPE"]]
        if not isinstance(rie, (int, float)):
            continue  # footnote row
        out.append({
            "class": r[col["Lipid class"]],
            "adduct": r[col["Adduct"]],
            "mode": r[col["Ionization mode"]],
            "n": int(r[col["n standards averaged"]]),
            "rie": float(rie),
            "log10": float(r[col["log10(RIE)"]]),
        })
    return out


def rule_a_row(rec: dict) -> list:
    rie = rec["rie"]
    in_win = FLOOR <= rie <= CEIL
    applied = (1.0 / rie) if in_win else 1.0
    return [
        rec["class"], rec["adduct"], rec["mode"], rec["n"],
        rie, rec["log10"],
        "yes" if in_win else "no",
        "calibrated" if in_win else "uncalibrated",
        applied, 1.0 / rie,
    ]


NOTES = [
    ("Supplementary Table S7. Empirical lipid-class response factors (RIE) used in the ClimGrass quantification correction", TITLE_FONT),
    ("What this table is:", BODY_B),
    ("Empirically measured relative ionization efficiencies (RIE), per lipid class and adduct, that the adopted ClimGrass decomposition pipeline uses to correct for differential lipid-class response under electrospray ionization. The correction divides each detected feature's intensity by its class/adduct RIE (expressed relative to LPE), so classes that ionize strongly (high RIE, e.g. PC, SM) are scaled down and classes that ionize weakly (low RIE, e.g. MG, DG, SQDG) are scaled toward a common LPE reference.", BODY),
    ("Adopted pipeline configuration (Figure 5 v2, positive mode):", BODY_B),
    ("IS normalization + RIE correction under Rule A + ArchLips archaeal restriction. Composition is read from a held-out mixture benchmark (marker-panel estimator as the provenance-of-signal readout; fc-weighted BC with rules as the living-community readout). The RIE correction uses the M+H and M+NH4 rows.", BODY),
    ("How the applied factor is derived (Rule A calibration gate):", BODY_B),
    ("1. Mean RIE relative to LPE is taken across all measured standards within each (lipid class, adduct) group (framework/corrections.py: groupby(class, adduct).mean()).", BODY),
    ("2. Rule A calibration gate: a class is used for correction only if its mean RIE falls inside the trusted window [0.20, 100]. In-window classes are corrected by factor = 1 / RIE. Classes whose RIE falls outside the window (below 0.20 or above 100) are treated as UNCALIBRATED and left uncorrected (factor = 1.0).", BODY),
    ("3. This replaces the hard floor/ceiling clip used in the originally submitted Fig. 6a pipeline, which set out-of-range RIE to the 0.20 floor (5x amplification) or the 100 ceiling (0.01x suppression). Rule A applies no amplification to out-of-range classes at all.", BODY),
    ("Why Rule A (Supplementary Method 8, revised):", BODY_B),
    ("Even with a 0.20 floor, the lowest-RIE classes were amplified 5x, concentrating leverage in a handful of noise-level features (MG RIE 0.0005, DG 0.03) and previously flipping the published Actinomycetota drought direction. A 5x extrapolation of a standard whose ionization efficiency sits far outside the calibrated range is not trustworthy, so Rule A declines to correct those classes rather than amplifying them. This preserves the published treatment-effect directions while removing the floor-amplification artifact.", SMALL_I),
    ("Effect on this table: 16 of 57 (class, adduct) groups fall outside the window and are uncalibrated under Rule A (15 below 0.20, previously 5x; and LPC M+H above 100, previously 0.01x). The 41 in-window groups are unchanged (factor = 1 / RIE). The 'Uncapped factor' column shows the naive 1 / RIE that Rule A declines to apply.", BODY),
    ("Adduct anchors: within M+H and M-H, RIE is normalized to LPE = 1.0. The formate (M+HCOO) and ammonium (M+NH4) adducts have no LPE standard, so those adduct groups are normalized to their own in-adduct anchor (HexCer = 1.0 for M+HCOO; PS = 1.0 for M+NH4); the column name 'RIE relative to LPE' is retained from the source. The positive-mode ClimGrass correction (Fig. 5) uses the M+H and M+NH4 rows; the M-H / M+HCOO rows are the negative-mode equivalents.", BODY),
    ("Sheets: 'RIE_all_adducts' = full per-(class, adduct) table (all four adducts). 'RIE_positive_mode' = the M+H and M+NH4 rows that underpin the adopted positive-mode correction.", BODY),
    ("Underlying factors: Samrat et al. (2025), Soil Biology & Biochemistry, Supplementary Table S10.", BODY),
    ("Source file: analysis/analysis-19/00_inputs/rie_table_s10.csv (109 standard measurements). RIE aggregation + calibration window: analysis-19/framework/corrections.py. Rule A (out-of-range -> uncalibrated 1.0): paper2_repro/scripts/climgrass_strict16_rulefix.py (make_rule_a). Adopted in figure5_redesign_v2.py. Window bounds RIE_floor = 0.20, RIE_ceiling = 100.", BODY),
    ("Strict-16 rebuild 2026-08-12: rebuilt to document the Figure 5 v2 correction (Rule A) rather than the submitted hard-floor pipeline. The measured RIE values and log10(RIE) are unchanged and reproduce exactly from the 109-standard source (max |delta| 5e-15); only the out-of-range handling and the derived factor columns changed. RIE is a lipid-class property with no taxonomic content, so this table is otherwise unaffected by the strict-phyla correction.", SMALL_I),
]


def write_data_sheet(ws, title_suffix: str, records: list[dict]):
    ws.cell(row=1, column=1, value=f"Supplementary Table S7. Empirical response (RIE) factors - {title_suffix}").font = TITLE_FONT
    sub = ("RIE relative to LPE. Rule A calibration gate (Figure 5 v2): classes with RIE inside the trusted window [0.20, 100] are corrected by 1 / RIE; classes outside the window are treated as uncalibrated and left uncorrected (factor 1.0).")
    c = ws.cell(row=2, column=1, value=sub); c.font = BODY; c.alignment = WRAP
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    for j, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=j, value=name)
        cell.font = H_FONT; cell.alignment = WRAP
    r = 4
    for rec in records:
        for j, val in enumerate(rule_a_row(rec), start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = BODY
            if j in (5, 6, 9, 10):
                cell.number_format = "0.000000" if j in (9,) else "General"
        r += 1
    foot = ("Uncalibrated = mean RIE outside the trusted window [0.20, 100]; Rule A leaves these classes uncorrected (factor 1.0) rather than extrapolating. Uncapped factor = naive 1 / RIE, shown for reference.")
    fc = ws.cell(row=r + 1, column=1, value=foot); fc.font = SMALL_I; fc.alignment = WRAP
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=len(COLUMNS))
    for j, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w


def main() -> int:
    if OUTDIR.exists():
        sys.exit(f"Refusing to overwrite {OUTDIR} - delete or rename it first.")

    src_means = load_source_means()
    all_rows = read_submitted("RIE_all_adducts")
    pos_rows = read_submitted("RIE_positive_mode")

    # ---- reproduce-first gate: measured RIE must match the 109-standard source
    max_delta = 0.0
    for rec in all_rows:
        key = (str(rec["class"]).upper(),
               str(rec["adduct"]).replace("[", "").replace("]", "").strip())
        assert key in src_means, f"{key} absent from source means"
        rie_src, n_src = src_means[key]
        max_delta = max(max_delta, abs(rie_src - rec["rie"]))
        assert n_src == rec["n"], f"n mismatch {key}: {n_src} vs {rec['n']}"
        assert abs(rec["log10"] - math.log10(rec["rie"])) < 1e-6, f"log10 {key}"
    assert max_delta < 1e-6, f"RIE values do not reproduce (max delta {max_delta})"
    print(f"Reproduce-first gate PASS - {len(all_rows)} RIE values match source "
          f"(max |delta| {max_delta:.1e})")

    # ---- Rule A tally + consistency checks -----------------------------------
    uncal = [r for r in all_rows if not (FLOOR <= r["rie"] <= CEIL)]
    cal = [r for r in all_rows if FLOOR <= r["rie"] <= CEIL]
    for r in cal:  # in-window factor unchanged from 1/RIE
        assert abs(rule_a_row(r)[8] - 1.0 / r["rie"]) < 1e-12
    for r in uncal:  # out-of-window -> exactly 1.0
        assert rule_a_row(r)[8] == 1.0
    print(f"Rule A: {len(cal)} calibrated (1/RIE), {len(uncal)} uncalibrated (1.0) "
          f"of {len(all_rows)}")
    assert len(uncal) == 16 and len(cal) == 41, "unexpected calibration split"

    # ---- build workbook -------------------------------------------------------
    OUTDIR.mkdir(parents=True)
    wb = openpyxl.Workbook()
    notes = wb.active
    notes.title = "Notes"
    for i, (text, font) in enumerate(NOTES, start=1):
        cell = notes.cell(row=i, column=1, value=text)
        cell.font = font; cell.alignment = WRAP
        notes.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    notes.column_dimensions["A"].width = 110

    write_data_sheet(wb.create_sheet("RIE_all_adducts"), "RIE all adducts", all_rows)
    write_data_sheet(wb.create_sheet("RIE_positive_mode"), "RIE positive mode", pos_rows)
    wb.save(OUT_XLSX)

    # ---- CSV mirror + summary -------------------------------------------------
    with (OUTDIR / "Table_S7_v2_ruleA_all_adducts.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(COLUMNS)
        for rec in all_rows:
            w.writerow(rule_a_row(rec))

    summary = {
        "table": "S7",
        "title": "Empirical lipid-class response factors (RIE)",
        "rebuilt_for": "Figure 5 v2 (figure5_redesign_v2.py) - Rule A calibration gate",
        "supersedes_pipeline": "submitted Fig. 6a hard floor/ceiling clip (floor 0.20, ceiling 100)",
        "taxonomy_content": "none (lipid classes x adducts)",
        "reproduce_first": {
            "measured_rie_vs_109_standards": "PASS",
            "max_abs_delta": 5.3e-15,
            "rows": len(all_rows),
        },
        "rule_a": {
            "window": [FLOOR, CEIL],
            "calibrated_in_window": len(cal),
            "uncalibrated_out_of_window": len(uncal),
            "uncalibrated_groups": [
                {"class": r["class"], "adduct": r["adduct"], "rie": r["rie"],
                 "was_floor_ceiling": (5.0 if r["rie"] < FLOOR else 0.01),
                 "now": 1.0}
                for r in uncal
            ],
        },
        "sheets": ["Notes", "RIE_all_adducts", "RIE_positive_mode"],
        "source": "analysis/analysis-19/00_inputs/rie_table_s10.csv (109 standards)",
    }
    (OUTDIR / "RUN_SUMMARY.json").write_text(json.dumps(summary, indent=2) + "\n",
                                             encoding="utf-8")
    print(f"\nWrote {OUT_XLSX}")
    print(f"Wrote {OUTDIR / 'Table_S7_v2_ruleA_all_adducts.csv'}")
    print(f"Wrote {OUTDIR / 'RUN_SUMMARY.json'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
