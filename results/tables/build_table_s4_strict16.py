#!/usr/bin/env python3
"""Supplementary Table S4 - distance-metric sensitivity of the lipidome-phylogeny
Mantel test, on the LOCKED release ncbi-phylum-2026-08-04-v1.

Reuses the Supplementary Figure 1d machinery unchanged (same profile builder,
same permutation test, same seeds) and sweeps three distance metrics in both
polarities.

Status: documented recomputation, not a reproduction. The published values used
the pre-correction unit scheme, and the phylogeny distance matrix for that
scheme is not in the repository - only `phylo_dist_16phyla.csv` exists - so the
published r values cannot be recomputed. Consistency check: this code path
returns Bray-Curtis r = 0.598 (POS) and 0.761 (NEG) against the published 0.603
and 0.767, i.e. the correction moves them by <0.01, which is reassuring but is
not a gate.

Improvement on the submitted table: it reported Jaccard and Cosine for positive
mode only, leaving negative mode as em-dashes. Both are computed here.

Jaccard is computed on presence/absence profiles (the standard for community
data); Bray-Curtis and cosine on the abundance profiles.
"""

from __future__ import annotations

import importlib.util
import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.spatial.distance import pdist, squareform

PROJECT_ROOT = Path(__file__).resolve().parents[2]
RELEASE_ROOT = PROJECT_ROOT / "outputs" / "analysis" / "ncbi-phylum-2026-08-04-v1"
TAX = RELEASE_ROOT / "taxonomy"
PANEL_D = RELEASE_ROOT / "suppfig1_full_strict16_2026-08-11_v1" / "panel_d_data"
PHYLO16 = PANEL_D / "phylo_dist_16phyla.csv"
SRC = PROJECT_ROOT / "paper2_repro" / "scripts" / "suppfig1_panel_d_permutations.py"
SUBMITTED = (PROJECT_ROOT / "outputs" / "tables"
             / "Table_S4_mantel_distance_metric_sensitivity.xlsx")
OUT_DIR = RELEASE_ROOT / "tables" / "table_s4_strict16_2026-08-12_v1"
OUT_XLSX = OUT_DIR / "Table_S4_mantel_distance_metric_sensitivity_strict16.xlsx"

METRICS = [
    ("Bray-Curtis", "braycurtis", "abundance"),
    ("Jaccard", "jaccard", "presence/absence"),
    ("Cosine", "cosine", "abundance"),
]
FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")


def load_src():
    spec = importlib.util.spec_from_file_location("panel_d", SRC)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


def main() -> int:
    if OUT_DIR.exists():
        sys.exit(f"Refusing to overwrite {OUT_DIR} - delete or rename it first.")
    if not PHYLO16.exists():
        sys.exit(f"missing {PHYLO16}; run suppfig1_full_strict16.py first")

    m = load_src()
    phylo = pd.read_csv(PHYLO16, index_col=0)
    summary = json.loads((TAX / "taxonomy_summary.json").read_text())
    units = sorted(summary["analysis_phyla"])

    neg_cols = pd.read_csv(m.NEG_META_S, nrows=1).columns
    neg_col = "sample_col" if "sample_col" in neg_cols else "original_header"

    rows = []
    for mode, table, meta, scol in [("POS", m.POS_TABLE, m.POS_META_S, "original_header"),
                                    ("NEG", m.NEG_TABLE, m.NEG_META_S, neg_col)]:
        q, cols, sp, sb = m.build_profiles(table, meta, scol, units)
        bp = sorted(set(sp.values()) & set(phylo.index))
        prof = np.array([q[[s for s in cols if sp[s] == p]].mean(axis=1).values
                         for p in bp])
        ph = phylo.loc[bp, bp].values
        for label, metric, basis in METRICS:
            x = (prof > 0).astype(bool) if metric == "jaccard" else prof
            d = squareform(pdist(x, metric=metric))
            # Pass the 'Overall Mantel' tag so mantel_perm takes the stable
            # SEEDS[mode]. Any other tag routes it through abs(hash(tag)), and
            # Python randomises string hashing per process, which makes the
            # permutation p values differ between runs. Using one seed per mode
            # also pairs the three metrics on the same permutation set.
            res = m.mantel_perm(d, ph, mode, "Overall Mantel")
            rows.append({
                "Distance metric": label,
                "Profile basis": basis,
                "Mode": mode,
                "n phyla": res["n_phyla"],
                "n pairs": res["n_pairs"],
                "Mantel r": round(float(res["observed_r"]), 3),
                "Mantel p (permutation)": float(res["mantel_p_greater"]),
                "Permutations": res.get("permutations", 9999),
                "Permutation method": res["permutation_method"],
                "Seed": res.get("seed"),
            })
            print(f"[{mode}] {label:<12} r={res['observed_r']:.3f} "
                  f"p={res['mantel_p_greater']:.4g}")

    long = pd.DataFrame(rows)
    # Column order mirrors the submitted table: metric, POS r, POS p, NEG r, NEG p.
    idx = long.set_index(["Distance metric", "Mode"])
    wide = pd.DataFrame({
        "Distance metric": [lbl for lbl, _, _ in METRICS],
        "POS Mantel r": [idx.loc[(l, "POS"), "Mantel r"] for l, _, _ in METRICS],
        "POS p-value": [idx.loc[(l, "POS"), "Mantel p (permutation)"]
                        for l, _, _ in METRICS],
        "NEG Mantel r": [idx.loc[(l, "NEG"), "Mantel r"] for l, _, _ in METRICS],
        "NEG p-value": [idx.loc[(l, "NEG"), "Mantel p (permutation)"]
                        for l, _, _ in METRICS],
    })

    sub = pd.read_excel(SUBMITTED, "Table", header=2)
    sub = sub[sub["Distance metric"].notna()]
    comp = wide.merge(sub, on="Distance metric", how="left",
                      suffixes=("", " (submitted)"))

    OUT_DIR.mkdir(parents=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as xw:
        wide.to_excel(xw, sheet_name="Table", index=False, startrow=3)
        long.to_excel(xw, sheet_name="Full_results", index=False, startrow=3)
        comp.to_excel(xw, sheet_name="Vs_submitted", index=False, startrow=3)

    wb = load_workbook(OUT_XLSX)
    notes = wb.create_sheet("Notes", 0)
    bc_pos = long[(long["Distance metric"] == "Bray-Curtis") & (long.Mode == "POS")]
    bc_neg = long[(long["Distance metric"] == "Bray-Curtis") & (long.Mode == "NEG")]
    lines = [
        ("Supplementary Table S4. Distance-metric sensitivity of the "
         "lipidome-phylogeny Mantel test (strict release)", True),
        ("", False),
        (f"Mantel r between the phylum-level lipidome distance matrix and the "
         f"phylogenetic distance matrix, over the {summary['n_analysis_phyla']} "
         f"analysis phyla (120 pairs), in both ionisation modes.", False),
        ("", False),
        ("Bray-Curtis is used for all main analyses; Jaccard and cosine are "
         "reported to show the phylogenetic signal is not an artefact of the "
         "distance metric.", False),
        ("", False),
        ("STATUS: documented recomputation, not a reproduction. The published "
         "values were computed on the pre-correction unit scheme, and the "
         "phylogeny distance matrix for that scheme is not in the repository "
         "(only the 16-phylum matrix exists), so they cannot be recomputed. "
         "Consistency check: this code path gives Bray-Curtis "
         f"r = {float(bc_pos['Mantel r'].iloc[0]):.3f} (POS) and "
         f"{float(bc_neg['Mantel r'].iloc[0]):.3f} (NEG) against the published "
         "0.603 and 0.767 - a shift below 0.01. Sheet 'Vs_submitted' places the "
         "two side by side.", False),
        ("", False),
        ("The submitted table reported Jaccard and cosine for positive mode "
         "only and left negative mode as em-dashes. Both are now computed.", False),
        ("", False),
        ("Jaccard is computed on presence/absence profiles, the standard for "
         "community data; Bray-Curtis and cosine on abundance profiles. The "
         "'Profile basis' column in 'Full_results' records this per row.", False),
        ("", False),
        ("p values are permutation-based (label permutations of the phylogeny "
         "matrix), not the parametric correlation p. Method and permutation "
         "count are given per row in 'Full_results'.", False),
        ("", False),
        ("Machinery is Supplementary Figure 1d's, unchanged: same profile "
         "builder (cross-batch features, detection >= 5%, mean intensity "
         ">= 500), same permutation test, same seeds.", False),
        ("", False),
        (f"Taxonomy release: {summary['taxonomy_release']} (locked). "
         "Producer: paper2_repro/scripts/build_table_s4_strict16.py", False),
    ]
    for i, (txt, bold) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=txt)
        c.font = Font(name=FONT, size=11, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    notes.column_dimensions["A"].width = 112

    titles = {
        "Table": ("Supplementary Table S4. Distance-metric sensitivity "
                  "(Mantel r, strict release)",
                  "Permutation p values. Negative mode now included for all "
                  "three metrics."),
        "Full_results": ("Supplementary Table S4b. Full Mantel results",
                         "Per metric and mode, with profile basis, pair counts "
                         "and permutation method."),
        "Vs_submitted": ("Supplementary Table S4c. Strict release vs submitted",
                         "Side-by-side. The submitted values are not "
                         "reproducible - see Notes."),
    }
    for sheet, (t1, t2) in titles.items():
        w = wb[sheet]
        w.cell(row=1, column=1, value=t1).font = Font(name=FONT, size=11, bold=True)
        w.cell(row=2, column=1, value=t2).font = Font(name=FONT, size=10, italic=True)
        for j in range(1, w.max_column + 1):
            hc = w.cell(row=4, column=j)
            hc.font = Font(name=FONT, size=10, bold=True)
            hc.fill = HDR_FILL
            hc.alignment = Alignment(wrap_text=True, vertical="bottom")
        for row in w.iter_rows(min_row=5, max_row=w.max_row, max_col=w.max_column):
            for c in row:
                c.font = Font(name=FONT, size=10)
        for j in range(1, w.max_column + 1):
            L = get_column_letter(j)
            longest = max((len(str(w.cell(row=i, column=j).value or ""))
                           for i in range(4, w.max_row + 1)), default=10)
            w.column_dimensions[L].width = min(max(longest + 2, 11), 34)
        w.freeze_panes = "A5"
    wb.save(OUT_XLSX)

    long.to_csv(OUT_DIR / "s4_mantel_metric_sensitivity_strict16.csv", index=False)
    (OUT_DIR / "RUN_SUMMARY.json").write_text(json.dumps({
        "taxonomy_release": summary["taxonomy_release"],
        "status": "documented recomputation; published values not reproducible "
                  "(legacy-scheme phylogeny matrix absent)",
        "n_phyla": int(long["n phyla"].iloc[0]),
        "n_pairs": int(long["n pairs"].iloc[0]),
        "results": long.to_dict("records"),
        "submitted": {"Bray-Curtis": {"POS": 0.603, "NEG": 0.767},
                      "Jaccard": {"POS": 0.361, "NEG": None},
                      "Cosine": {"POS": 0.396, "NEG": None}},
        "improvement": "negative mode computed for Jaccard and cosine, which "
                       "the submitted table left blank",
    }, indent=2) + "\n", encoding="utf-8")

    print()
    print(wide.to_string(index=False))
    print(f"\nWrote {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
