#!/usr/bin/env python3
"""Supplementary Table S1 - sample inventory reconciled to the LOCKED release
ncbi-phylum-2026-08-04-v1.

The submitted workbook (outputs/tables/Table_S1_sample_inventory.xlsx) carries
the pre-correction phylum labels. This rebuild joins every row to the release
taxonomy and republishes it with:

  * Phylum      -> the release's ncbi_phylum (strict assignment)
  * Kingdom     -> the release's ecological_group (policy naming:
                   Viridiplantae / Protists, not Plantae / Protozoa)
  * the submitted labels retained alongside, so the change is auditable
  * NCBI phylum taxid, taxonomy scope, and release id per row

Join: exact `Original Code` -> release `sample_name`, then a case- and
hyphen-insensitive fallback for the remainder. All 170 submitted rows resolve.
Acquisition timestamps are NOT stripped: `I8POS` and `I8POS_20240925010944` are
distinct injections present in both sources.

Two defects in the submitted table are reported rather than silently fixed:
  1. `OE25F3POS.mzML` appears twice, so "170 biological samples" is 169 unique;
  2. the per-phylum `n_species` column is not reproducible and is internally
     inconsistent (the same problem that forced Supplementary Figure 7 onto
     genus counts), so S1b reports genera and states why.
"""

from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
RELEASE_ROOT = PROJECT_ROOT / "outputs" / "analysis" / "ncbi-phylum-2026-08-04-v1"
TAX = RELEASE_ROOT / "taxonomy"
SUBMITTED = PROJECT_ROOT / "outputs" / "tables" / "Table_S1_sample_inventory.xlsx"
OUT_DIR = RELEASE_ROOT / "tables" / "table_s1_strict16_2026-08-12_v1"
OUT_XLSX = OUT_DIR / "Table_S1_sample_inventory_strict16.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
CHANGED_FILL = PatternFill("solid", fgColor="FFF2CC")


def resolve_join(sub: pd.DataFrame, rel: pd.DataFrame) -> pd.Series:
    exact = {s: s for s in rel["sample_name"]}
    loose = {str(s).upper().replace("-", ""): s for s in rel["sample_name"]}

    def r(x):
        x = str(x).strip()
        if x in exact:
            return exact[x]
        return loose.get(x.upper().replace("-", ""))

    return sub["Original Code"].map(r)


def main() -> int:
    if OUT_DIR.exists():
        sys.exit(f"Refusing to overwrite {OUT_DIR} - delete or rename it first.")

    policy = json.loads((PROJECT_ROOT / "paper2_repro" / "config"
                         / "taxonomy_policy.json").read_text(encoding="utf-8"))
    assert policy["status"] == "locked"
    summary = json.loads((TAX / "taxonomy_summary.json").read_text())
    taxids = policy["verified_ncbi_phyla"]

    sub = pd.read_excel(SUBMITTED, "Per_sample_inventory", header=3)
    sub = sub[sub["GNPS Filename"].notna()].copy()
    rel = pd.read_csv(TAX / "sample_metadata_POS_ncbi_phylum.csv")

    sub["_match"] = resolve_join(sub, rel)
    if sub["_match"].isna().any():
        sys.exit("unresolved rows in the join - inspect before rebuilding")

    dup_mask = sub.duplicated(subset=["GNPS Filename", "Original Code"], keep=False)
    n_dup_rows = int(dup_mask.sum())

    # Six core-candidate samples of the release are absent from the submitted
    # inventory. An inventory table that omits profiled organisms would put S1b
    # in conflict with the release and with Supplementary Figure 7, so they are
    # added and flagged. Culture-condition fields are unavailable for them.
    covered = set(sub["_match"])
    missing = rel[(rel["taxonomy_scope"] == "core_candidate")
                  & (~rel["sample_name"].isin(covered))].copy()
    add = pd.DataFrame({
        "GNPS Filename": missing["sample_name"].astype(str) + ".mzML",
        "Original Code": missing["sample_name"],
        "FBMN_Batch": missing["batch"],
        "Kingdom": missing["kingdom"],
        "Phylum": missing["source_phylum"],
        "Class": missing.get("class"), "Order": missing.get("order"),
        "Family": missing.get("family"), "Genus": missing.get("genus"),
        "Species": missing.get("species"),
        "_match": missing["sample_name"],
        "_added": "added from release (absent from submitted table)",
    })
    n_added = len(add)
    sub["_added"] = ""
    # Keep the duplicated row visible for audit but mark it so it is excluded
    # from the per-phylum counts; otherwise S1b would report Ascomycota 28
    # against the release's 27 and disagree with Supplementary Figure 7.
    second = sub.duplicated(subset=["GNPS Filename", "Original Code"], keep="first")
    sub.loc[second, "_added"] = "duplicate row in submitted table (excluded from counts)"
    sub = pd.concat([sub, add], ignore_index=True)

    r = rel.set_index("sample_name")
    sub["Phylum (release)"] = sub["_match"].map(r["ncbi_phylum"])
    sub["Kingdom (release)"] = sub["_match"].map(r["ecological_group"])
    sub["Taxonomy scope"] = sub["_match"].map(r["taxonomy_scope"])
    sub["NCBI phylum taxid"] = sub["Phylum (release)"].map(taxids)
    sub["Taxonomy release"] = summary["taxonomy_release"]
    sub["Phylum (as submitted)"] = sub["Phylum"]
    sub["Kingdom (as submitted)"] = sub["Kingdom"]
    sub["Row source"] = sub["_added"].replace("", "submitted table")
    sub["Count in summary"] = (
        (sub["Taxonomy scope"] == "core_candidate")
        & (~sub["_added"].str.startswith("duplicate"))).map({True: "yes", False: "no"})
    sub["Label changed"] = (sub["Phylum (release)"].fillna("")
                            != sub["Phylum (as submitted)"].fillna("")).map(
                                {True: "yes", False: ""})

    ordered = ["GNPS Filename", "Original Code", "FBMN_Batch",
               "Kingdom (release)", "Phylum (release)", "NCBI phylum taxid",
               "Taxonomy scope", "Kingdom (as submitted)", "Phylum (as submitted)",
               "Label changed", "Class", "Order", "Family", "Genus", "Species",
               "NCBI ID", "KEGG ID", "Lipid Recovery %", "Media_DSMZ_or_type",
               "Temperature_C", "Oxygen_conditions", "pH", "Taxonomy release",
               "Row source", "Count in summary"]
    s1a = sub[ordered].copy()

    core = s1a[s1a["Count in summary"] == "yes"].copy()
    core["_g"] = core["Genus"].astype(str).str.strip().str.lower().str.split().str[0]
    grp = (core.groupby(["Kingdom (release)", "Phylum (release)"])
           .agg(n_samples=("Original Code", "size"), n_genera=("_g", "nunique"))
           .reset_index()
           .sort_values(["Kingdom (release)", "Phylum (release)"]))
    grp["In strict analysis set"] = grp["Phylum (release)"].isin(
        summary["analysis_phyla"]).map({True: "yes", False: "no (n<2)"})

    OUT_DIR.mkdir(parents=True)
    shutil.copy2(SUBMITTED, OUT_DIR / "Table_S1_sample_inventory_SUBMITTED.xlsx")

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as xw:
        s1a.to_excel(xw, "Per_sample_inventory", index=False, startrow=3)
        grp.to_excel(xw, "Per_phylum_summary", index=False, startrow=3)

    wb = load_workbook(OUT_XLSX)

    notes = wb.create_sheet("Notes", 0)
    changed = int((s1a["Label changed"] == "yes").sum())
    from_sub = s1a["Row source"] == "submitted table"
    n_from_submitted = int(from_sub.sum())
    n_unique_submitted = int(s1a.loc[from_sub, "Original Code"].nunique())
    lines = [
        ("Supplementary Table S1. Per-sample and per-phylum inventory "
         "(strict release)", True),
        ("", False),
        (f"Taxonomy release: {summary['taxonomy_release']} (locked). Phylum and "
         "kingdom are taken from that release; the labels as submitted are "
         "retained alongside so every change is auditable.", False),
        (f"Rows: {n_from_submitted} carried over from the submitted workbook "
         f"plus {n_added} added from the release = {len(s1a)}. NOTE: "
         f"OE25F3POS.mzML appears twice in the submitted workbook "
         f"({n_dup_rows} rows share that identity), so its stated count of 170 "
         f"biological samples is {n_unique_submitted} unique samples.", False),
        (f"{n_added} core-candidate samples of the release are ABSENT from the "
         "submitted inventory and have been added here, flagged in the "
         "'Row source' column: ACNeffE25-123POS (Discosea, Acanthamoeba), "
         "Lu19POS (Cyanobacteriota, lichen), OE11-3-E4-P27POS (Streptophyta, "
         "Ginkgo), OE21-4-F3POS (Ascomycota, Tuber), OE25102891POS "
         "(Actinomycetota, Arthrobacter), OE25DSM1988POS (Ascomycota, "
         "Aspergillus). Culture-condition fields are unavailable for them. "
         "Without these rows S1b would disagree with the release and with "
         "Supplementary Figure 7.", False),
        (f"Phylum label changed for {changed} of {len(s1a)} rows.", False),
        ("", False),
        ("Relabelling applied (release authority, not edited here): "
         "Amoebozoa split to Discosea / Evosea; Euryarchaeota and Halobacteriota "
         "to Methanobacteriota; Crenarchaeota to Thermoproteota; Bryophyta, "
         "Marchantiophyta, Tracheophyta, Magnoliophyta and Charophyta to "
         "Streptophyta; Mortierellomycota separated from Mucoromycota.", False),
        ("", False),
        ("Kingdom uses the locked policy's ecological_group naming "
         "(Viridiplantae, Protists), which differs from the submitted table's "
         "Plantae / Protozoa. Ecological groups are display summaries and are "
         "never analysis units.", False),
        ("", False),
        (f"Counts: {summary['n_collection_phyla']} collection phyla; "
         f"{summary['n_analysis_phyla']} analysis phyla (n>=2 in both "
         f"polarities). Below threshold: "
         f"{', '.join(summary['below_threshold'])}.", False),
        ("Taxonomy scope: core_candidate rows have a valid NCBI phylum ancestor. "
         "descriptive_only rows (Bicosoecida, Rootnodules, Mixed) and viral "
         "samples are retained in the inventory but excluded from phylum "
         "inference, per policy.", False),
        ("", False),
        ("The submitted per-phylum n_species column is NOT reproduced. It is not "
         "derivable from any metadata column present and is internally "
         "inconsistent (for example Euryarchaeota was published as 4 species "
         "across 9 distinct genera). S1b therefore reports distinct genera, the "
         "finest rank the metadata determines for every sample - the same "
         "decision taken for Supplementary Figure 7.", False),
        ("", False),
        ("Scope: positive mode, matching the submitted table. The release also "
         "covers 195 negative-mode core samples.", False),
        ("", False),
        ("Producer: paper2_repro/scripts/build_table_s1_strict16.py", False),
    ]
    for i, (txt, bold) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=txt)
        c.font = Font(name=FONT, size=11, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    notes.column_dimensions["A"].width = 118

    titles = {
        "Per_sample_inventory": (
            f"Supplementary Table S1a. Per-sample inventory "
            f"({s1a['Original Code'].nunique()} unique biological samples, "
            f"{len(s1a)} rows as submitted)",
            "Phylum and kingdom from the locked release; submitted labels retained."),
        "Per_phylum_summary": (
            "Supplementary Table S1b. Per-phylum inventory (core samples)",
            "Sample and genus counts per release phylum. n_species is not "
            "reported - see Notes."),
    }
    for sheet, (t1, t2) in titles.items():
        ws = wb[sheet]
        ws.cell(row=1, column=1, value=t1).font = Font(name=FONT, size=11, bold=True)
        ws.cell(row=2, column=1, value=t2).font = Font(name=FONT, size=10, italic=True)
        ncol = ws.max_column
        for j in range(1, ncol + 1):
            h = ws.cell(row=4, column=j)
            h.font = Font(name=FONT, size=10, bold=True)
            h.fill = HDR_FILL
            h.alignment = Alignment(wrap_text=True, vertical="bottom")
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=ncol):
            for c in row:
                c.font = Font(name=FONT, size=10)
        for j in range(1, ncol + 1):
            letter = get_column_letter(j)
            longest = max((len(str(ws.cell(row=i, column=j).value or ""))
                           for i in range(4, min(ws.max_row, 300) + 1)), default=10)
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 42)
        ws.freeze_panes = "A5"

    ws = wb["Per_sample_inventory"]
    hdr = {ws.cell(row=4, column=j).value: j for j in range(1, ws.max_column + 1)}
    jchg, jphy = hdr["Label changed"], hdr["Phylum (release)"]
    for i in range(5, ws.max_row + 1):
        if ws.cell(row=i, column=jchg).value == "yes":
            ws.cell(row=i, column=jphy).fill = CHANGED_FILL

    # n_samples as a live formula against S1a so the summary recalculates.
    ws2 = wb["Per_phylum_summary"]
    h2 = {ws2.cell(row=4, column=j).value: j for j in range(1, ws2.max_column + 1)}
    col_phy = get_column_letter(jphy)
    last = ws.max_row
    for i in range(5, ws2.max_row + 1):
        pcell = f"{get_column_letter(h2['Phylum (release)'])}{i}"
        ws2.cell(row=i, column=h2["n_samples"]).value = (
            f"=COUNTIFS(Per_sample_inventory!${col_phy}$5:${col_phy}${last},{pcell},"
            f"Per_sample_inventory!${get_column_letter(hdr['Count in summary'])}$5:"
            f"${get_column_letter(hdr['Count in summary'])}${last},\"yes\")")
    ws2.cell(row=ws2.max_row + 2, column=1,
             value=("n_samples is a live COUNTIFS against S1a. n_genera is computed "
                    "from the Genus column (first token, case-normalised) and written "
                    "as a value - a distinct count has no formula form that "
                    "evaluates reliably outside Excel.")
             ).font = Font(name=FONT, size=9, italic=True)

    wb.save(OUT_XLSX)

    grp.to_csv(OUT_DIR / "per_phylum_summary_strict16.csv", index=False)
    s1a.to_csv(OUT_DIR / "per_sample_inventory_strict16.csv", index=False)
    (OUT_DIR / "RUN_SUMMARY.json").write_text(json.dumps({
        "taxonomy_release": summary["taxonomy_release"],
        "rows_as_submitted": int(len(s1a)),
        "unique_samples": int(s1a["Original Code"].nunique()),
        "duplicated_rows_in_submitted": n_dup_rows,
        "core_samples_added_from_release": n_added,
        "phylum_label_changed_rows": changed,
        "core_candidate_rows": int((s1a["Taxonomy scope"] == "core_candidate").sum()),
        "collection_phyla": summary["n_collection_phyla"],
        "analysis_phyla": summary["n_analysis_phyla"],
        "n_species_column": "not reproduced - not derivable and internally "
                            "inconsistent; genera reported instead",
        "scope": "positive mode, matching the submitted table",
    }, indent=2) + "\n", encoding="utf-8")

    print(f"rows {len(s1a)} ({s1a['Original Code'].nunique()} unique), "
          f"phylum changed on {changed}")
    print(grp.to_string(index=False))
    print(f"\nWrote {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
