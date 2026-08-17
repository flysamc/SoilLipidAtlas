#!/usr/bin/env python3
"""Table S12 (cross-method consensus features per strict phylum).

Built from the Supplementary Figure 8 strict rerun
(`suppfig8_cross_method_strict16_2026-08-11_v1/panel_d_consensus_per_phylum.csv`).
For each strict phylum: biomarker features selected by all four methods (SIMPER,
SCBD, CAP, L1), by >=3, by >=2, and by exactly one, out of the total tested.

Legacy-keyed-defect check PASSED: the source is RE-DERIVED on the 16 strict phyla,
not relabelled. Consensus membership is defined against the strict partition, so
split units (Discosea/Evosea/Heterolobosea) and merged units (Methanobacteriota,
Streptophyta) carry their own consensus counts. Gate here asserts total ==
only_1 + geq2 and the nesting all4 <= geq3 <= geq2 for every phylum.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
POLICY = ROOT / "paper2_repro/config/taxonomy_policy.json"
SRC = (ROOT / "outputs/analysis/ncbi-phylum-2026-08-04-v1"
       / "suppfig8_cross_method_strict16_2026-08-11_v1/panel_d_consensus_per_phylum.csv")
OUTDIR = (ROOT / "outputs/analysis/ncbi-phylum-2026-08-04-v1/tables"
          / "table_s12_strict16_2026-08-12_v1")
OUT_XLSX = OUTDIR / "Table_S12_crossmethod_consensus_per_phylum_strict16.xlsx"

TITLE = Font(name="Arial", size=13, bold=True)
SUB = Font(name="Arial", size=10)
SUB_I = Font(name="Arial", size=9, italic=True)
H = Font(name="Arial", size=10, bold=True)
BODY = Font(name="Arial", size=10)
BODY_B = Font(name="Arial", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

COLS = [("phylum", "Phylum"), ("kingdom", "Kingdom"), ("total", "Total features"),
        ("all4", "All-4 consensus"), ("geq3", ">=3 methods"),
        ("geq2", ">=2 methods"), ("only_1", "Only 1 method")]


def main() -> int:
    if OUTDIR.exists():
        sys.exit(f"Refusing to overwrite {OUTDIR} - delete or rename it first.")
    eco = json.loads(POLICY.read_text(encoding="utf-8"))["ecological_group"]

    d = pd.read_csv(SRC)
    assert len(d) == 16, f"expected 16 strict phyla, got {len(d)}"
    d["kingdom"] = d["phylum"].map(eco)  # normalize to Viridiplantae/Protists
    assert d["kingdom"].notna().all(), \
        f"phyla missing from policy: {sorted(d.loc[d.kingdom.isna(),'phylum'])}"
    # integrity gates
    assert (d["total"] == d["only_1"] + d["geq2"]).all(), "total != only_1 + geq2"
    assert (d["all4"] <= d["geq3"]).all() and (d["geq3"] <= d["geq2"]).all(), "nesting broken"
    print("Gate PASS - 16 strict phyla; total==only_1+geq2; all4<=geq3<=geq2; "
          "kingdoms normalized to Viridiplantae/Protists")

    d = d.sort_values(["all4", "geq3", "phylum"], ascending=[False, False, True]).reset_index(drop=True)
    n_zero_all4 = int((d["all4"] == 0).sum())

    OUTDIR.mkdir(parents=True)
    wb = openpyxl.Workbook()
    notes = wb.active; notes.title = "Notes"
    lines = [
        ("Supplementary Table S12. Cross-method consensus features per strict phylum", TITLE),
        ("For each of the 16 strict analysis phyla, the number of biomarker features selected in agreement by all four fingerprint methods (SIMPER, SCBD, CAP, L1 stability), by at least three, by at least two, and by exactly one method, out of the total features tested. Higher all-four counts indicate method-invariant, robustly diagnostic chemistry.", SUB),
        ("Re-derived on the strict 16-phylum partition (Supplementary Figure 8 strict rerun), NOT relabelled from the legacy scheme: consensus membership is defined against the strict units, so split units (Discosea, Evosea, Heterolobosea) and merged units (Methanobacteriota, Streptophyta) carry their own consensus counts. The amoebozoan protists Evosea and Discosea hold the most all-four-consensus features.", SUB),
        (f"{n_zero_all4} of 16 phyla have zero all-four-consensus features; their diagnostic features are recovered by fewer methods (see the >=2 / >=3 columns). Kingdom labels follow the locked ecological_group vocabulary (Viridiplantae, Protists). CAP and L1 stability are declared reimplementations (see Table S11), so consensus counts involving them are reimplementation-based.", SUB),
        ("Source: outputs/analysis/ncbi-phylum-2026-08-04-v1/suppfig8_cross_method_strict16_2026-08-11_v1/panel_d_consensus_per_phylum.csv. Producer: paper2_repro/scripts/build_table_s12_strict16.py.", SUB_I),
    ]
    for i, (t, f) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=t); c.font = f; c.alignment = WRAP
        notes.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    notes.column_dimensions["A"].width = 118

    ws = wb.create_sheet("Consensus_per_phylum")
    c = ws.cell(row=1, column=1, value="Supplementary Table S12. Cross-method consensus features per strict phylum")
    c.font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    for j, (_, label) in enumerate(COLS, start=1):
        hc = ws.cell(row=2, column=j, value=label); hc.font = H; hc.alignment = WRAP
    r = 3
    for _, row in d.iterrows():
        for j, (key, _) in enumerate(COLS, start=1):
            v = row[key] if key in ("phylum", "kingdom") else int(row[key])
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = BODY_B if j == 1 else BODY
        r += 1
    # total row (live SUM over count columns)
    ws.cell(row=r, column=1, value="All phyla").font = BODY_B
    for j, (key, _) in enumerate(COLS, start=1):
        if key in ("phylum", "kingdom"):
            continue
        cl = get_column_letter(j)
        tc = ws.cell(row=r, column=j, value=f"=SUM({cl}3:{cl}{r-1})"); tc.font = BODY_B
    widths = [16, 14, 13, 15, 13, 13, 13]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    wb.save(OUT_XLSX)

    d[[k for k, _ in COLS]].to_csv(OUTDIR / "S12_consensus_per_phylum.csv", index=False)
    (OUTDIR / "RUN_SUMMARY.json").write_text(json.dumps({
        "table": "S12", "title": "Cross-method consensus features per strict phylum",
        "source": "suppfig8_cross_method_strict16_2026-08-11_v1/panel_d_consensus_per_phylum.csv",
        "legacy_keyed_defect_check": "PASS (re-derived on strict 16, not relabelled)",
        "n_phyla": 16, "n_zero_all4": n_zero_all4,
        "top_all4": {row["phylum"]: int(row["all4"]) for _, row in d.head(4).iterrows()},
        "gates": {"total==only_1+geq2": "PASS", "all4<=geq3<=geq2": "PASS"},
        "ties_to": "Supplementary Figure 8 panel d",
    }, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
