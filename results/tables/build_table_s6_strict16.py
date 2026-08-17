#!/usr/bin/env python3
"""Table S6 (expected composition ranges) on the LOCKED release naming.

S6 is the one supplementary table with nothing to re-derive: its ranges are
order-of-magnitude expectations synthesised from the soil-biology literature
(Bar-On 2018, Joergensen & Wichern 2008, Bates 2011, Jackson 1996, ...), not
computed from our biomarker set. It is organism-group level, not phylum level,
so the legacy-keyed-artifact defect does not apply -- there is no target/
background contrast to recompute when phyla merge.

The ONLY change required for the strict-16 release is to conform the two
organism-group labels to the locked `ecological_group` vocabulary in
config/taxonomy_policy.json:

    Plantae  -> Viridiplantae
    Protozoa -> Protists

Everything else (ranges, bases, references, formatting, the entire Notes sheet)
is carried over verbatim from the frozen submitted workbook. We therefore edit
the frozen file in place rather than reconstruct it, which guarantees byte-level
fidelity except for the two intended relabels plus one provenance line.

Gates:
  * reproduce-first: the two source labels must be exactly 'Plantae' / 'Protozoa'
    at the known addresses before we touch anything;
  * exactly two pre-existing cells change, and nothing else;
  * the final six group labels must all be members of the locked
    ecological_group value set.
"""

from __future__ import annotations

import csv
import json
import shutil
import sys
from copy import copy
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[2]
POLICY = PROJECT_ROOT / "paper2_repro" / "config" / "taxonomy_policy.json"
SRC = PROJECT_ROOT / "outputs" / "tables" / "Table_S6_expected_composition_ranges.xlsx"
RELEASE = PROJECT_ROOT / "outputs" / "analysis" / "ncbi-phylum-2026-08-04-v1"
OUTDIR = RELEASE / "tables" / "table_s6_strict16_2026-08-12_v1"
OUT_XLSX = OUTDIR / "Table_S6_expected_composition_ranges_strict16.xlsx"

RELABEL = {"Plantae": "Viridiplantae", "Protozoa": "Protists"}
# reproduce-first anchors: where the labels are known to live in the frozen file
ANCHORS = {"A6": "Plantae", "A9": "Protozoa"}


def snapshot(wb) -> dict:
    out = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    out[(ws.title, c.coordinate)] = c.value
    return out


def main() -> int:
    if OUTDIR.exists():
        sys.exit(f"Refusing to overwrite {OUTDIR} - delete or rename it first.")

    policy = json.loads(POLICY.read_text(encoding="utf-8"))
    assert policy.get("status") == "locked", "policy is not locked"
    eco_vocab = set(policy["ecological_group"].values())
    print(f"Locked ecological_group vocabulary: {sorted(eco_vocab)}")
    for tgt in RELABEL.values():
        assert tgt in eco_vocab, f"target label {tgt!r} not in locked vocabulary"

    wb = openpyxl.load_workbook(SRC)
    tbl = wb["Table"]

    # ---- reproduce-first gate -------------------------------------------------
    for addr, expected in ANCHORS.items():
        got = tbl[addr].value
        assert got == expected, (
            f"reproduce-first FAIL: Table!{addr} is {got!r}, expected {expected!r}. "
            "The frozen submitted S6 has changed; stop and investigate."
        )
    before = snapshot(wb)
    print("Reproduce-first gate PASS - frozen labels found at A6/A9.")

    # ---- apply the two relabels ----------------------------------------------
    for addr, expected in ANCHORS.items():
        tbl[addr] = RELABEL[expected]

    # final group vocabulary must be a subset of the locked ecological groups
    groups = [tbl[f"A{r}"].value for r in range(4, 10)]
    print(f"Final organism-group labels: {groups}")
    bad = [g for g in groups if g not in eco_vocab]
    assert not bad, f"group labels not in locked vocabulary: {bad}"

    # ---- provenance line into the Notes sheet --------------------------------
    notes = wb["Notes"]
    last = max((c.row for row in notes.iter_rows() for c in row
                if c.value not in (None, "")), default=1)
    note_row = last + 2
    prov = (
        "Strict-16 update 2026-08-12: organism-group labels conformed to the "
        "locked ecological_group vocabulary in taxonomy_policy.json (release "
        "ncbi-phylum-2026-08-04-v1): Plantae -> Viridiplantae, Protozoa -> "
        "Protists. Ranges, bases, and references are unchanged from the frozen "
        "2026-06-09 version; these expectations are literature-derived and are "
        "not affected by the taxonomy correction."
    )
    cell = notes.cell(row=note_row, column=1, value=prov)
    # match the caveat/frozen styling (Arial 9 italic) and the A:H merge
    template = notes["A12"]
    cell.font = copy(template.font)
    cell.alignment = copy(template.alignment)
    notes.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)

    # ---- verify: exactly two pre-existing cells changed, nothing lost --------
    after = snapshot(wb)
    changed = {k for k in before if before[k] != after.get(k)}
    dropped = {k for k in before if k not in after}
    added = {k for k in after if k not in before}
    assert dropped == set(), f"cells lost: {dropped}"
    assert changed == {("Table", "A6"), ("Table", "A9")}, (
        f"unexpected changes: {changed}"
    )
    assert added == {("Notes", f"A{note_row}")}, f"unexpected additions: {added}"
    print(f"Verify PASS - changed exactly {sorted(changed)}; "
          f"added provenance at Notes!A{note_row}; nothing dropped.")

    # ---- write outputs --------------------------------------------------------
    OUTDIR.mkdir(parents=True)
    wb.save(OUT_XLSX)

    # CSV mirror of the Table sheet (for diffable audit)
    with (OUTDIR / "Table_S6_strict16.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        for row in tbl.iter_rows(values_only=True):
            r = list(row)
            while r and r[-1] is None:
                r.pop()
            if any(v is not None for v in r):
                w.writerow(r)

    summary = {
        "table": "S6",
        "title": "Literature-reported expected composition ranges by organism group",
        "source_frozen": str(SRC.relative_to(PROJECT_ROOT)),
        "release": "ncbi-phylum-2026-08-04-v1",
        "rederivation_required": False,
        "reason_no_rederivation": (
            "organism-group-level literature expectations; not derived from the "
            "biomarker set, no target/background contrast, so no legacy-keyed defect"
        ),
        "relabels_applied": RELABEL,
        "cells_changed": ["Table!A6", "Table!A9"],
        "provenance_note_cell": f"Notes!A{note_row}",
        "final_groups": groups,
        "ranges_bases_refs": "unchanged (verbatim)",
        "gates": {
            "reproduce_first": "PASS",
            "exactly_two_changes": "PASS",
            "labels_in_locked_vocab": "PASS",
        },
    }
    (OUTDIR / "RUN_SUMMARY.json").write_text(json.dumps(summary, indent=2) + "\n",
                                             encoding="utf-8")
    print(f"\nWrote {OUT_XLSX}")
    print(f"Wrote {OUTDIR / 'Table_S6_strict16.csv'}")
    print(f"Wrote {OUTDIR / 'RUN_SUMMARY.json'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
