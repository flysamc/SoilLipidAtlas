#!/usr/bin/env python3
"""Table S14 (ClimGrass verified-soil feature overlap) on strict 16 phyla.

Built from the Supplementary Figure 8 strict rerun
(`suppfig8_cross_method_strict16_2026-08-11_v1/panel_c_climgrass_overlap.csv`).
For each fingerprint method and top-K level, the fraction of the 696 spectrally
verified ClimGrass soil features captured within that method's top-K biomarker
set (and the reverse fraction + raw overlap). Global metric over the strict
top-K rankings; no per-phylum keying, so the legacy-keyed defect does not apply.
CAP and L1 stability carry the reimplementation caveats noted in Tables S11/S13.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
SRC = (ROOT / "outputs/analysis/ncbi-phylum-2026-08-04-v1"
       / "suppfig8_cross_method_strict16_2026-08-11_v1/panel_c_climgrass_overlap.csv")
OUTDIR = (ROOT / "outputs/analysis/ncbi-phylum-2026-08-04-v1/tables"
          / "table_s14_strict16_2026-08-12_v1")
OUT_XLSX = OUTDIR / "Table_S14_climgrass_overlap_fractions_strict16.xlsx"

METHOD_LABEL = {"simper": "SIMPER", "scbd": "SCBD", "cap": "CAP", "stability_l1": "L1 stability"}
METHOD_ORDER = ["simper", "scbd", "cap", "stability_l1"]
K_ORDER = [100, 250, 500, 1000, 2500]

TITLE = Font(name="Arial", size=13, bold=True)
SUB = Font(name="Arial", size=10)
SUB_I = Font(name="Arial", size=9, italic=True)
H = Font(name="Arial", size=10, bold=True)
BODY = Font(name="Arial", size=10)
BODY_B = Font(name="Arial", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def main() -> int:
    if OUTDIR.exists():
        sys.exit(f"Refusing to overwrite {OUTDIR} - delete or rename it first.")
    d = pd.read_csv(SRC)
    piv = d.pivot(index="method", columns="K", values="frac_of_climgrass_in_topK")
    assert set(piv.index) == set(METHOD_ORDER) and list(piv.columns) == K_ORDER
    n_cg = int(d["n_climgrass_features"].iloc[0])
    assert (d["n_climgrass_features"] == n_cg).all()
    print(f"Gate PASS - method x K present; ClimGrass verified-soil features = {n_cg}")

    OUTDIR.mkdir(parents=True)
    wb = openpyxl.Workbook()
    notes = wb.active; notes.title = "Notes"
    lines = [
        ("Supplementary Table S14. ClimGrass verified-soil feature overlap (strict 16 phyla)", TITLE),
        (f"Fraction of the {n_cg} spectrally verified ClimGrass soil features captured within each method's top-K biomarker set, at each K level (Supplementary Figure 8 strict rerun). Higher fractions mean the method's top-ranked atlas biomarkers better cover the features actually detected in ClimGrass soil. Uses the corrected {n_cg}-feature verified-soil substrate.", SUB),
        ("The 'Overlap_detail' sheet also gives the reverse fraction (fraction of each method's top-K set that is ClimGrass-detected) and the raw overlap counts. Variance-based methods (SCBD, SIMPER) capture the ClimGrass set fastest; CAP and L1 stability are supervised and cover it more slowly.", SUB),
        ("CAP and L1 stability are declared reimplementations (see Tables S11 and S13); overlap fractions involving them are reimplementation-based. This is a global top-K overlap metric (no per-phylum keying).", SUB),
        ("Source: outputs/analysis/ncbi-phylum-2026-08-04-v1/suppfig8_cross_method_strict16_2026-08-11_v1/panel_c_climgrass_overlap.csv. Producer: paper2_repro/scripts/build_table_s14_strict16.py.", SUB_I),
    ]
    for i, (t, f) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=t); c.font = f; c.alignment = WRAP
        notes.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    notes.column_dimensions["A"].width = 118

    # --- S14a: fraction of ClimGrass captured, method x K ---
    ws = wb.create_sheet("ClimGrass_overlap")
    c = ws.cell(row=1, column=1, value=f"Supplementary Table S14a. Fraction of the {n_cg} verified ClimGrass features captured in each method's top-K")
    c.font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(K_ORDER))
    for j, name in enumerate(["Method"] + [f"K = {k}" for k in K_ORDER], start=1):
        hc = ws.cell(row=2, column=j, value=name); hc.font = H; hc.alignment = WRAP
    r = 3
    for m in METHOD_ORDER:
        ws.cell(row=r, column=1, value=METHOD_LABEL[m]).font = BODY_B
        for j, k in enumerate(K_ORDER, start=2):
            cell = ws.cell(row=r, column=j, value=round(float(piv.loc[m, k]), 4))
            cell.font = BODY; cell.number_format = "0.0%"
        r += 1
    ws.column_dimensions["A"].width = 16
    for j in range(2, 2 + len(K_ORDER)):
        ws.column_dimensions[get_column_letter(j)].width = 10

    # --- S14b: detail ---
    ws2 = wb.create_sheet("Overlap_detail")
    c = ws2.cell(row=1, column=1, value="Supplementary Table S14b. Overlap detail (counts and both directions)")
    c.font = TITLE
    cols = ["Method", "K", "Top-K features", "ClimGrass features", "Overlap",
            "Frac. of ClimGrass in top-K", "Frac. of top-K in ClimGrass"]
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    for j, name in enumerate(cols, start=1):
        hc = ws2.cell(row=2, column=j, value=name); hc.font = H; hc.alignment = WRAP
    r = 3
    for m in METHOD_ORDER:
        for k in K_ORDER:
            row = d[(d["method"] == m) & (d["K"] == k)].iloc[0]
            vals = [METHOD_LABEL[m], k, int(row["n_features"]), int(row["n_climgrass_features"]),
                    int(row["n_overlap"]), round(float(row["frac_of_climgrass_in_topK"]), 4),
                    round(float(row["frac_of_topK_in_climgrass"]), 4)]
            for j, v in enumerate(vals, start=1):
                cell = ws2.cell(row=r, column=j, value=v)
                cell.font = BODY_B if j == 1 else BODY
                if j in (6, 7):
                    cell.number_format = "0.0%"
            r += 1
    for j, w in enumerate([15, 7, 14, 15, 9, 20, 20], start=1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    wb.save(OUT_XLSX)

    d.to_csv(OUTDIR / "S14_climgrass_overlap.csv", index=False)
    (OUTDIR / "RUN_SUMMARY.json").write_text(json.dumps({
        "table": "S14", "title": "ClimGrass verified-soil feature overlap",
        "source": "suppfig8_cross_method_strict16_2026-08-11_v1/panel_c_climgrass_overlap.csv",
        "n_climgrass_features": n_cg,
        "metric": "frac_of_climgrass_in_topK (method x K), + reverse fraction + counts",
        "legacy_keyed_defect_check": "N/A (global top-K overlap, not per-phylum)",
        "ties_to": "Supplementary Figure 8 panel c",
    }, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
