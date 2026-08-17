#!/usr/bin/env python3
"""Table S13 (LOO classification accuracy) on strict 16 phyla.

Primary view = leave-one-out 1-nearest-phylum accuracy by method x top-K, from the
Supplementary Figure 8 strict rerun (`panel_b_loo_accuracy.csv`). Secondary view =
per-phylum accuracy at the full substrate, from the Supplementary Figure 4 strict
rerun (`suppfig4_loo_strict16_2026-08-11_v1/loo_per_unit_strict16.csv`), which is the
64.6% (106/164) NNLS classifier behind Supp Fig 4.

DECLARED REIMPLEMENTATION: the exact submitted classifier/producer (52.8%) is
unrecovered, so these absolute accuracies are a named NNLS/method reimplementation,
not a reproduction. Do not relabel the submitted values onto this.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
POLICY = ROOT / "paper2_repro/config/taxonomy_policy.json"
RE = ROOT / "outputs/analysis/ncbi-phylum-2026-08-04-v1"
PANELB = RE / "suppfig8_cross_method_strict16_2026-08-11_v1/panel_b_loo_accuracy.csv"
PERUNIT = RE / "suppfig4_loo_strict16_2026-08-11_v1/loo_per_unit_strict16.csv"
OUTDIR = RE / "tables/table_s13_strict16_2026-08-12_v1"
OUT_XLSX = OUTDIR / "Table_S13_loo_classification_accuracy_strict16.xlsx"

METHOD_LABEL = {"simper": "SIMPER", "scbd": "SCBD", "cap": "CAP", "stability_l1": "L1 stability"}
METHOD_ORDER = ["simper", "scbd", "cap", "stability_l1"]
K_ORDER = [100, 250, 500, 1000, 2500]

TITLE = Font(name="Arial", size=13, bold=True)
SUB = Font(name="Arial", size=10)
SUB_I = Font(name="Arial", size=9, italic=True)
H = Font(name="Arial", size=10, bold=True)
BODY = Font(name="Arial", size=10)
BODY_B = Font(name="Arial", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def main() -> int:
    if OUTDIR.exists():
        sys.exit(f"Refusing to overwrite {OUTDIR} - delete or rename it first.")
    eco = json.loads(POLICY.read_text(encoding="utf-8"))["ecological_group"]

    pb = pd.read_csv(PANELB)
    pu = pd.read_csv(PERUNIT)
    piv = pb.pivot(index="method", columns="K", values="accuracy")
    assert set(piv.index) == set(METHOD_ORDER) and list(piv.columns) == K_ORDER
    assert (pb["n_total"] == 164).all(), "n_total should be 164"
    baseline = float(pb["full_substrate_baseline"].iloc[0])
    # per-unit consistency: overall == baseline == 106/164
    assert pu["correct"].sum() == 106 and pu["n"].sum() == 164, \
        f"per-unit totals {pu['correct'].sum()}/{pu['n'].sum()} != 106/164"
    assert abs(pu["correct"].sum() / pu["n"].sum() - baseline) < 1e-3, "baseline mismatch"
    pu["kingdom"] = pu["unit"].map(eco)
    assert pu["kingdom"].notna().all(), "unit missing from policy"
    print(f"Gate PASS - method x K present; n_total=164; full-substrate baseline "
          f"{baseline:.3f} == per-unit 106/164; kingdoms normalized")

    pu = pu.sort_values(["accuracy_pct", "unit"], ascending=[False, True]).reset_index(drop=True)

    OUTDIR.mkdir(parents=True)
    wb = openpyxl.Workbook()
    notes = wb.active; notes.title = "Notes"
    lines = [
        ("Supplementary Table S13. Leave-one-out classification accuracy (strict 16 phyla)", TITLE),
        ("Leave-one-out 1-nearest-phylum classification accuracy over the 164 strict-labelled samples. 'Accuracy_by_method_K' gives accuracy for each fingerprint method (SIMPER, SCBD, CAP, L1 stability) at each top-K feature level (Supplementary Figure 8 strict rerun). 'Accuracy_by_phylum' gives per-phylum accuracy at the full substrate - the 64.6% (106/164) 16-phylum NNLS classifier behind Supplementary Figure 4.", SUB),
        ("Full-substrate baseline (all features, 16-phylum NNLS): 64.6% (106/164). Variance-based methods (SIMPER, SCBD) match the baseline from moderate K; the supervised L1 stability exceeds it; CAP (a bounded reimplementation) underperforms.", SUB),
        ("DECLARED REIMPLEMENTATION: the exact submitted classifier/producer (reported 52.8%) is unrecovered, so these absolute accuracies are a named NNLS/method reimplementation on the strict 16-phylum label set, NOT a reproduction of the submitted values. CAP and L1 stability additionally carry the reimplementation caveats noted in Table S11. Kingdom labels follow the locked ecological_group vocabulary (Viridiplantae, Protists).", SUB),
        ("Source: outputs/analysis/ncbi-phylum-2026-08-04-v1/suppfig8_cross_method_strict16_2026-08-11_v1/panel_b_loo_accuracy.csv and suppfig4_loo_strict16_2026-08-11_v1/loo_per_unit_strict16.csv. Producer: paper2_repro/scripts/build_table_s13_strict16.py.", SUB_I),
    ]
    for i, (t, f) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=t); c.font = f; c.alignment = WRAP
        notes.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    notes.column_dimensions["A"].width = 118

    # --- S13a: method x K ---
    ws = wb.create_sheet("Accuracy_by_method_K")
    c = ws.cell(row=1, column=1, value="Supplementary Table S13a. LOO accuracy by method and top-K")
    c.font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(K_ORDER))
    for j, name in enumerate(["Method"] + [f"K = {k}" for k in K_ORDER], start=1):
        hc = ws.cell(row=2, column=j, value=name); hc.font = H; hc.alignment = WRAP
    r = 3
    for m in METHOD_ORDER:
        ws.cell(row=r, column=1, value=METHOD_LABEL[m]).font = BODY_B
        for j, k in enumerate(K_ORDER, start=2):
            cell = ws.cell(row=r, column=j, value=round(float(piv.loc[m, k]), 4))
            cell.font = BODY; cell.number_format = "0.0%"
        r += 1
    bl = ws.cell(row=r + 1, column=1,
                 value=f"Full-substrate baseline (all features, 16-phylum NNLS): "
                       f"{baseline*100:.1f}% (106/164)")
    bl.font = SUB_I
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=1 + len(K_ORDER))
    ws.column_dimensions["A"].width = 16
    for j in range(2, 2 + len(K_ORDER)):
        ws.column_dimensions[get_column_letter(j)].width = 10

    # --- S13b: per-phylum at full substrate ---
    ws2 = wb.create_sheet("Accuracy_by_phylum")
    c = ws2.cell(row=1, column=1, value="Supplementary Table S13b. Per-phylum LOO accuracy at the full substrate (64.6% overall)")
    c.font = TITLE
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    for j, name in enumerate(["Phylum", "Kingdom", "Correct", "n samples", "Accuracy"], start=1):
        hc = ws2.cell(row=2, column=j, value=name); hc.font = H; hc.alignment = WRAP
    r = 3
    for _, row in pu.iterrows():
        ws2.cell(row=r, column=1, value=row["unit"]).font = BODY_B
        ws2.cell(row=r, column=2, value=row["kingdom"]).font = BODY
        ws2.cell(row=r, column=3, value=int(row["correct"])).font = BODY
        ws2.cell(row=r, column=4, value=int(row["n"])).font = BODY
        ac = ws2.cell(row=r, column=5, value=f"=C{r}/D{r}"); ac.font = BODY; ac.number_format = "0.0%"
        r += 1
    ws2.cell(row=r, column=1, value="All phyla").font = BODY_B
    ws2.cell(row=r, column=3, value=f"=SUM(C3:C{r-1})").font = BODY_B
    ws2.cell(row=r, column=4, value=f"=SUM(D3:D{r-1})").font = BODY_B
    tc = ws2.cell(row=r, column=5, value=f"=C{r}/D{r}"); tc.font = BODY_B; tc.number_format = "0.0%"
    for j, w in enumerate([16, 14, 9, 11, 10], start=1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    wb.save(OUT_XLSX)

    pb.to_csv(OUTDIR / "S13a_accuracy_by_method_K.csv", index=False)
    pu.to_csv(OUTDIR / "S13b_accuracy_by_phylum.csv", index=False)
    (OUTDIR / "RUN_SUMMARY.json").write_text(json.dumps({
        "table": "S13", "title": "LOO classification accuracy",
        "primary": "method x K (suppfig8 panel_b_loo_accuracy.csv)",
        "secondary": "per-phylum at full substrate (suppfig4 loo_per_unit_strict16.csv)",
        "full_substrate_baseline": round(baseline, 4), "n_samples": 164, "n_correct": 106,
        "status": "DECLARED REIMPLEMENTATION (submitted 52.8% producer unrecovered)",
        "legacy_keyed_defect_check": "N/A (sample classification re-derived on strict labels)",
        "ties_to": ["Supplementary Figure 8 panel b", "Supplementary Figure 4 (64.6%)"],
    }, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
