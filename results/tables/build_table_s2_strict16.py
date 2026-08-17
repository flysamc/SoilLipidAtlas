#!/usr/bin/env python3
"""Supplementary Table S2 - FBMN batch composition, rebuilt on the LOCKED
release ncbi-phylum-2026-08-04-v1.

Reproduce-first gate: the six per-batch feature counts are recomputed from the
FBMN quantification tables and must match the submitted values exactly
(40,936 / 84,347 / 89,731 / 43,206 / 20,424 / 60,234; total 338,878).

What changes:
  * sample counts were approximations ("~20" ... total "~205"). They are now
    exact, and split into all injections in the batch versus biological
    core-candidate samples, because the approximations conflated the two.
  * "Primary kingdoms" mixed ranks and used retired labels (Bryophyta,
    Cyanobacteria, Amoebozoa, and the phyla Bacillota / Actinomycetota listed
    as kingdoms). It is replaced by two rank-explicit columns derived from the
    release: organism groups and strict phyla, each with counts.
  * the batch-kingdom confounding note is restated against the strict labels.

A second sheet gives the full batch x phylum matrix, which is what a reviewer
needs to judge the confounding for themselves.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
RELEASE_ROOT = PROJECT_ROOT / "outputs" / "analysis" / "ncbi-phylum-2026-08-04-v1"
TAX = RELEASE_ROOT / "taxonomy"
SUBMITTED = PROJECT_ROOT / "outputs" / "tables" / "Table_S2_fbmn_batch_composition.xlsx"
FBMN_POS = (PROJECT_ROOT / "external" / "SOILMASS_PRODUCER_RECOVERY_2026-08-04"
            / "payload" / "core" / "workspace" / "analysis" / "FBMN_all_batches_POS")
OUT_DIR = RELEASE_ROOT / "tables" / "table_s2_strict16_2026-08-12_v1"
OUT_XLSX = OUT_DIR / "Table_S2_fbmn_batch_composition_strict16.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")


def count_features(batch_dir: Path) -> int:
    q = list((batch_dir / "results").glob("*quantification_table.csv"))
    if not q:
        sys.exit(f"no quantification table in {batch_dir}")
    with open(q[0], encoding="utf-8", errors="replace") as fh:
        return sum(1 for _ in fh) - 1


def main() -> int:
    if OUT_DIR.exists():
        sys.exit(f"Refusing to overwrite {OUT_DIR} - delete or rename it first.")

    summary = json.loads((TAX / "taxonomy_summary.json").read_text())
    sub = pd.read_excel(SUBMITTED, "Table", header=2)
    sub = sub[sub["Batch"].notna() & (sub["Batch"].astype(str) != "Total")].copy()
    sub["Batch"] = sub["Batch"].astype(str).str.zfill(2)
    pub_features = dict(zip(sub["Batch"], sub["Features"].astype(int)))
    task_ids = dict(zip(sub["Batch"], sub["GNPS2 task ID"]))
    pub_names = dict(zip(sub["Batch"], sub["Name"]))

    print("STAGE 1 - reproduce-first gate on feature counts")
    feats, dirs = {}, {}
    for d in sorted(FBMN_POS.glob("batch_0*")):
        b = d.name.split("_")[1]
        feats[b] = count_features(d)
        dirs[b] = d.name
        ok = feats[b] == pub_features.get(b)
        print(f"  batch {b}: {feats[b]:>7,}  published {pub_features.get(b):>7,}  "
              f"{'MATCH' if ok else 'DIFFERS'}")
        if not ok:
            sys.exit("GATE FAIL - feature counts do not reproduce; stopping.")
    if sum(feats.values()) != int(sub["Features"].astype(int).sum()):
        sys.exit("GATE FAIL - feature total does not reproduce")
    print(f"  GATE PASS (total {sum(feats.values()):,})")

    rel = pd.read_csv(TAX / "sample_metadata_POS_ncbi_phylum.csv")
    core = rel[rel["taxonomy_scope"] == "core_candidate"]

    rows = []
    for b in sorted(feats):
        dname = dirs[b]
        allb = rel[rel["batch"] == dname]
        cb = core[core["batch"] == dname]
        g = cb["ecological_group"].value_counts()
        p = cb["ncbi_phylum"].value_counts()
        rows.append({
            "Batch": b,
            "Name": pub_names.get(b),
            "Samples (all injections)": int(len(allb)),
            "Samples (biological)": int(len(cb)),
            "Features": feats[b],
            "Organism groups (n)": ", ".join(f"{k} {v}" for k, v in g.items()),
            "Phyla, strict release (n)": ", ".join(f"{k} {v}" for k, v in p.items()),
            "GNPS2 task ID": task_ids.get(b),
        })
    t = pd.DataFrame(rows)

    mat = (pd.crosstab(core["batch"], core["ncbi_phylum"])
           .reindex([dirs[b] for b in sorted(feats)]))
    mat.index.name = "Batch"
    mat = mat.reset_index()

    OUT_DIR.mkdir(parents=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as xw:
        t.to_excel(xw, sheet_name="Table", index=False, startrow=3)
        mat.to_excel(xw, sheet_name="Batch_by_phylum", index=False, startrow=3)

    wb = load_workbook(OUT_XLSX)

    # totals as live formulas
    ws = wb["Table"]
    h = {ws.cell(row=4, column=j).value: j for j in range(1, ws.max_column + 1)}
    last = ws.max_row
    trow = last + 1
    ws.cell(row=trow, column=1, value="Total").font = Font(name=FONT, size=10, bold=True)
    for col in ("Samples (all injections)", "Samples (biological)", "Features"):
        L = get_column_letter(h[col])
        c = ws.cell(row=trow, column=h[col], value=f"=SUM({L}5:{L}{last})")
        c.font = Font(name=FONT, size=10, bold=True)

    n_core = int(len(core))
    ani = core[core["ecological_group"] == "Animalia"]["batch"].value_counts()
    pro = core[core["ecological_group"] == "Protists"]["batch"].value_counts()
    b06 = core[core["batch"] == dirs["06"]]["ecological_group"].value_counts()

    notes = wb.create_sheet("Notes", 0)
    lines = [
        ("Supplementary Table S2. FBMN batch composition (strict release)", True),
        ("", False),
        ("Per-batch sample counts, feature counts, organism composition and "
         "GNPS2 task IDs. Positive mode.", False),
        ("", False),
        (f"Feature counts are unchanged from the submitted table and were "
         f"reproduced exactly from the FBMN quantification tables "
         f"(total {sum(feats.values()):,}).", False),
        ("", False),
        ("Sample counts were approximations in the submitted table (\"~20\" ... "
         "total \"~205\"). They are now exact. Two columns are given because the "
         "approximations conflated them: all injections assigned to the batch, "
         f"and biological samples with a valid NCBI phylum ancestor "
         f"({n_core} across the six batches).", False),
        ("", False),
        ("The submitted \"Primary kingdoms\" column mixed taxonomic ranks and "
         "used retired labels - Bryophyta, Cyanobacteria, Amoebozoa, and the "
         "phyla Bacillota and Actinomycetota listed as kingdoms. It is replaced "
         "by two rank-explicit columns taken from the locked release. Organism "
         "groups are display summaries; phyla are the analysis units.", False),
        ("", False),
        ("Batch-organism confounding, restated on the strict labels:", True),
        (f"  Animalia is confined to a single batch: all "
         f"{int(ani.sum())} animal samples are in {ani.index[0]}.", False),
        (f"  Protists occur in only two batches: "
         f"{', '.join(f'{k} ({v})' for k, v in pro.items())}.", False),
        (f"  {dirs['06']} is taxonomically homogeneous: "
         f"{', '.join(f'{k} {v}' for k, v in b06.items())} - it contains no "
         "eukaryotes at all.", False),
        ("  Viral samples (n=2) were acquired in batch 06 but are excluded from "
         "the study scope by author instruction, so they do not appear in the "
         "counts above.", False),
        ("", False),
        ("Sheet 'Batch_by_phylum' gives the full batch x phylum matrix so the "
         "confounding can be assessed directly.", False),
        ("", False),
        (f"Taxonomy release: {summary['taxonomy_release']} (locked). "
         "Producer: paper2_repro/scripts/build_table_s2_strict16.py", False),
    ]
    for i, (txt, bold) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=txt)
        c.font = Font(name=FONT, size=11, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    notes.column_dimensions["A"].width = 112

    titles = {
        "Table": ("Supplementary Table S2. FBMN batch composition",
                  "Sample counts exact; feature counts reproduced from the "
                  "quantification tables."),
        "Batch_by_phylum": ("Supplementary Table S2b. Batch x phylum matrix "
                            "(biological samples)",
                            "Counts of core-candidate samples per strict phylum."),
    }
    for sheet, (t1, t2) in titles.items():
        w = wb[sheet]
        w.cell(row=1, column=1, value=t1).font = Font(name=FONT, size=11, bold=True)
        w.cell(row=2, column=1, value=t2).font = Font(name=FONT, size=10, italic=True)
        for j in range(1, w.max_column + 1):
            hc = w.cell(row=4, column=j)
            hc.font = Font(name=FONT, size=10, bold=True)
            hc.fill = HDR_FILL
            hc.alignment = Alignment(wrap_text=True, vertical="bottom")
        for row in w.iter_rows(min_row=5, max_row=w.max_row, max_col=w.max_column):
            for c in row:
                if not c.font.bold:
                    c.font = Font(name=FONT, size=10)
                c.alignment = Alignment(wrap_text=True, vertical="top")
        for j in range(1, w.max_column + 1):
            L = get_column_letter(j)
            longest = max((len(str(w.cell(row=i, column=j).value or ""))
                           for i in range(4, w.max_row + 1)), default=10)
            w.column_dimensions[L].width = min(max(longest + 2, 10), 46)
        w.freeze_panes = "A5"

    wb.save(OUT_XLSX)

    t.to_csv(OUT_DIR / "s2_batch_composition_strict16.csv", index=False)
    mat.to_csv(OUT_DIR / "s2_batch_by_phylum.csv", index=False)
    (OUT_DIR / "RUN_SUMMARY.json").write_text(json.dumps({
        "taxonomy_release": summary["taxonomy_release"],
        "feature_count_gate": "PASS - all six batches and the total reproduce exactly",
        "features_total": int(sum(feats.values())),
        "samples_all_injections": int(t["Samples (all injections)"].sum()),
        "samples_biological": n_core,
        "submitted_sample_counts": "approximate (~20 ... total ~205); now exact",
        "primary_kingdoms_column": "replaced by rank-explicit group and phylum "
                                   "columns; retired labels removed",
        "confounding": {"Animalia_batches": ani.to_dict(),
                        "Protists_batches": pro.to_dict(),
                        "batch_06_groups": b06.to_dict()},
    }, indent=2) + "\n", encoding="utf-8")

    print()
    print(t[["Batch", "Name", "Samples (all injections)", "Samples (biological)",
             "Features"]].to_string(index=False))
    print(f"\nWrote {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
