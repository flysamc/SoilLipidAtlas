#!/usr/bin/env python3
"""Table S11 (cross-method dendrogram-reconstruction Mantel r) on strict 16 phyla.

Built from the Supplementary Figure 8 strict rerun
(`suppfig8_cross_method_strict16_2026-08-11_v1/panel_a_mantel_curves.csv` +
`panel_a_random_null.csv`). For each method and top-K feature level, the K
top-ranked features rebuild the phylum dendrogram and its cophenetic distances
are Mantel-correlated against the full-feature dendrogram (r->1 = the reduced
fingerprint reconstructs the full topology).

SIMPER and SCBD are exact strict recomputations; CAP is a bounded best-fit
reimplementation and L1 stability a declared reimplementation (exact historical
configs unrecovered), so CAP/L1 absolute values are reimplementations.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
S8 = (ROOT / "outputs/analysis/ncbi-phylum-2026-08-04-v1"
      / "suppfig8_cross_method_strict16_2026-08-11_v1")
OUTDIR = (ROOT / "outputs/analysis/ncbi-phylum-2026-08-04-v1/tables"
          / "table_s11_strict16_2026-08-12_v1")
OUT_XLSX = OUTDIR / "Table_S11_crossmethod_dendrogram_mantel_strict16.xlsx"

METHOD_LABEL = {"simper": "SIMPER", "scbd": "SCBD", "cap": "CAP",
                "stability_l1": "L1 stability"}
METHOD_ORDER = ["simper", "scbd", "cap", "stability_l1"]
K_ORDER = [100, 250, 500, 1000, 2500]

TITLE = Font(name="Arial", size=13, bold=True)
SUB = Font(name="Arial", size=10)
SUB_I = Font(name="Arial", size=9, italic=True)
H = Font(name="Arial", size=10, bold=True)
BODY = Font(name="Arial", size=10)
BODY_B = Font(name="Arial", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def main() -> int:
    if OUTDIR.exists():
        sys.exit(f"Refusing to overwrite {OUTDIR} - delete or rename it first.")

    curves = pd.read_csv(S8 / "panel_a_mantel_curves.csv")
    null = pd.read_csv(S8 / "panel_a_random_null.csv")
    piv = curves.pivot(index="method", columns="K", values="mantel_r_vs_full")
    assert set(piv.index) == set(METHOD_ORDER), f"methods: {list(piv.index)}"
    assert list(piv.columns) == K_ORDER, f"K columns: {list(piv.columns)}"
    null_by_k = null.set_index("K_per_phylum")
    print("Gate PASS - 4 methods x 5 K present; null has all K")

    OUTDIR.mkdir(parents=True)
    wb = openpyxl.Workbook()

    # ---- Notes ----
    notes = wb.active; notes.title = "Notes"
    lines = [
        ("Supplementary Table S11. Cross-method fingerprint validation - dendrogram-reconstruction Mantel r (strict 16 phyla)", TITLE),
        ("Per-method dendrogram-reconstruction Mantel r at each top-K feature level, computed on the strict 16-phylum lipid distance matrices (Supplementary Figure 8 strict rerun). For each method and K, the K top-ranked features rebuild the phylum dendrogram and its cophenetic distances are Mantel-correlated against the full-feature dendrogram; r approaching 1 means the reduced fingerprint reconstructs the full topology.", SUB),
        ("Random-K-features null: for each K, features are drawn at random and the reconstruction Mantel r recomputed. The 'Mantel_reconstruction' sheet lists the null mean and 95th percentile beneath the methods; the 'Random_null' sheet gives mean, SD and 5th/95th percentiles with the approximate union feature count. A method beats the null when its r sits above the null 95th percentile.", SUB),
        ("SIMPER and SCBD are exact strict recomputations. CAP is a bounded best-fit reimplementation and L1 stability a declared reimplementation (their exact historical configurations are unrecovered), so CAP and L1 absolute values are reimplementations, not reproductions.", SUB),
        ("This within-manuscript Supplementary Table S11 (cross-method validation) is distinct from the external 'Samrat et al. 2025, Supplementary Table S10' (RIE factors) that underpins Table S7.", SUB),
        ("Source: outputs/analysis/ncbi-phylum-2026-08-04-v1/suppfig8_cross_method_strict16_2026-08-11_v1/ (panel_a_mantel_curves.csv, panel_a_random_null.csv). Producer: paper2_repro/scripts/build_table_s11_strict16.py.", SUB_I),
    ]
    for i, (t, f) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=t); c.font = f; c.alignment = WRAP
        notes.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    notes.column_dimensions["A"].width = 118

    # ---- Mantel_reconstruction (method x K + null rows) ----
    ws = wb.create_sheet("Mantel_reconstruction")
    c = ws.cell(row=1, column=1, value="Supplementary Table S11a. Dendrogram-reconstruction Mantel r by method and top-K")
    c.font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(K_ORDER))
    hdr = ["Method"] + [f"K = {k}" for k in K_ORDER]
    for j, name in enumerate(hdr, start=1):
        hc = ws.cell(row=2, column=j, value=name); hc.font = H; hc.alignment = WRAP
    r = 3
    for m in METHOD_ORDER:
        ws.cell(row=r, column=1, value=METHOD_LABEL[m]).font = BODY_B
        for j, k in enumerate(K_ORDER, start=2):
            cell = ws.cell(row=r, column=j, value=round(float(piv.loc[m, k]), 3))
            cell.font = BODY; cell.number_format = "0.000"
        r += 1
    # null comparison rows
    for label, col in [("Random-K null (mean)", "null_mean"), ("Random-K null (95th pct)", "null_q95")]:
        ws.cell(row=r, column=1, value=label).font = SUB_I
        for j, k in enumerate(K_ORDER, start=2):
            cell = ws.cell(row=r, column=j, value=round(float(null_by_k.loc[k, col]), 3))
            cell.font = SUB_I; cell.number_format = "0.000"
        r += 1
    ws.column_dimensions["A"].width = 22
    for j in range(2, 2 + len(K_ORDER)):
        ws.column_dimensions[get_column_letter(j)].width = 10

    # ---- Random_null (full) ----
    ws2 = wb.create_sheet("Random_null")
    c = ws2.cell(row=1, column=1, value="Supplementary Table S11b. Random-K-features null distribution")
    c.font = TITLE
    ncol = 6
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    cols = ["K (per phylum)", "approx. K (union)", "null mean", "null SD", "null 5th pct", "null 95th pct"]
    for j, name in enumerate(cols, start=1):
        hc = ws2.cell(row=2, column=j, value=name); hc.font = H; hc.alignment = WRAP
    for i, k in enumerate(K_ORDER, start=3):
        row = null_by_k.loc[k]
        vals = [k, int(row["approx_K_union"]), round(float(row["null_mean"]), 3),
                round(float(row["null_sd"]), 3), round(float(row["null_q05"]), 3),
                round(float(row["null_q95"]), 3)]
        for j, v in enumerate(vals, start=1):
            cell = ws2.cell(row=i, column=j, value=v); cell.font = BODY
            if j >= 3:
                cell.number_format = "0.000"
    for j, w in enumerate([14, 16, 10, 10, 12, 12], start=1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    wb.save(OUT_XLSX)

    # CSV + summary
    piv.reindex(METHOD_ORDER).rename(index=METHOD_LABEL).to_csv(OUTDIR / "S11_mantel_by_method.csv")
    (OUTDIR / "RUN_SUMMARY.json").write_text(json.dumps({
        "table": "S11", "title": "Cross-method dendrogram-reconstruction Mantel r",
        "source": "suppfig8_cross_method_strict16_2026-08-11_v1 (panel_a_mantel_curves.csv, panel_a_random_null.csv)",
        "methods": {METHOD_LABEL[m]: {f"K{k}": round(float(piv.loc[m, k]), 3) for k in K_ORDER} for m in METHOD_ORDER},
        "provenance": "SIMPER/SCBD exact strict; CAP bounded reimplementation; L1 declared reimplementation",
        "ties_to": "Supplementary Figure 8 (strict rerun) panel a",
    }, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
