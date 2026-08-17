#!/usr/bin/env python3
"""Table S9 (SIRIUS / CANOPUS / CSI:FingerID / MSNovelist coverage) rebuilt
per strict phylum, both polarities.

Built from the integrated strict annotation evidence
(annotation/annotation_evidence_{pos,neg}.csv), which merges the LISC
SIRIUS/CANOPUS runs (initial atlas jobs 5800445 POS, 5802297 NEG; 2026-08-13
gap-fill jobs 5862224 and 5869613) with the historical full-atlas SIRIUS cache
via build_strict_annotation_evidence.py. These are the
same files Figure 2, Table S5 and Supplementary Figure 7 are built from, so S9
is consistent with them by construction.

Sheets:
  Coverage_overall       per-pipeline-step hit counts + rates, POS and NEG
  Coverage_by_phylum_POS  per strict phylum coverage (counts + live-formula rates)
  Coverage_by_phylum_NEG
  NPC_pathway_by_phylum_POS  NPClassifier pathway x phylum matrix (CANOPUS)
  NPC_pathway_by_phylum_NEG

Denominator = all strict biomarkers per phylum. SIRIUS needs usable MS2 and
precursor m/z <= 850 (a hard SIRIUS limit), so 'usable MS2' is shown as the
eligible pool that bounds the achievable formula rate.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
POLICY = ROOT / "paper2_repro" / "config" / "taxonomy_policy.json"
RELEASE = ROOT / "outputs" / "analysis" / "ncbi-phylum-2026-08-04-v1"
ANN = RELEASE / "annotation"
OUTDIR = RELEASE / "tables" / "table_s9_strict16_2026-08-13_v2"
OUT_XLSX = OUTDIR / "Table_S9_annotation_coverage_by_phylum_strict16.xlsx"

KING_ORDER = ["Bacteria", "Archaea", "Fungi", "Viridiplantae", "Animalia", "Protists"]
# pipeline coverage metrics: (label, evidence column)
METRICS = [
    ("Usable MS2", "has_usable_ms2"),
    ("SIRIUS molecular formula", "has_sirius_formula"),
    ("CANOPUS compound class", "has_canopus_class"),
    ("CSI:FingerID structure", "has_csi_fingerid_structure"),
    ("MSNovelist de novo structure", "has_denovo_structure"),
    ("DreaMS spectral neighbour", "has_dreams_result"),
]
NPC_ORDER = ["Fatty acids", "Terpenoids", "Polyketides", "Amino acids and Peptides",
             "Alkaloids", "Carbohydrates", "Shikimates and Phenylpropanoids"]

TITLE = Font(name="Arial", size=13, bold=True)
SUB = Font(name="Arial", size=10)
SUB_I = Font(name="Arial", size=9, italic=True)
H = Font(name="Arial", size=10, bold=True)
BODY = Font(name="Arial", size=10)
BODY_B = Font(name="Arial", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def as_bool(s: pd.Series) -> pd.Series:
    if s.dtype == bool:
        return s.fillna(False)
    if s.dtype == object:
        return s.astype(str).str.strip().str.lower().isin(["true", "1", "1.0", "yes"])
    return s.fillna(0).astype(float) > 0


def load(mode: str, eco: dict) -> pd.DataFrame:
    f = ANN / f"annotation_evidence_{mode.lower()}.csv"
    d = pd.read_csv(f, low_memory=False)
    d["kingdom"] = d["phylum"].map(eco)  # normalize to policy (Viridiplantae/Protists)
    assert d["kingdom"].notna().all(), \
        f"{mode}: phyla missing from policy: {sorted(set(d.loc[d.kingdom.isna(),'phylum']))}"
    for _, col in METRICS:
        d[col] = as_bool(d[col])
    return d


def per_phylum(d: pd.DataFrame) -> pd.DataFrame:
    g = d.groupby("phylum")
    rows = []
    for phy, sub in g:
        row = {"phylum": phy, "kingdom": sub["kingdom"].iloc[0], "n": len(sub)}
        for label, col in METRICS:
            row[col] = int(sub[col].sum())
        rows.append(row)
    df = pd.DataFrame(rows)
    df["_k"] = df["kingdom"].map(lambda k: KING_ORDER.index(k))
    return df.sort_values(["_k", "phylum"]).drop(columns="_k").reset_index(drop=True)


def npc_matrix(d: pd.DataFrame, phyla_order: list[str]) -> pd.DataFrame:
    sub = d[as_bool(d["has_canopus_class"])]
    m = (sub.pivot_table(index="phylum", columns="NPC#pathway", values="feature_id",
                         aggfunc="count", fill_value=0)
         .reindex(index=phyla_order, columns=NPC_ORDER, fill_value=0).astype(int))
    return m


def main() -> int:
    if OUTDIR.exists():
        sys.exit(f"Refusing to overwrite {OUTDIR} - delete or rename it first.")
    eco = json.loads(POLICY.read_text(encoding="utf-8"))["ecological_group"]

    pos = load("POS", eco)
    neg = load("NEG", eco)
    pp = {"POS": per_phylum(pos), "NEG": per_phylum(neg)}
    order = {m: pp[m]["phylum"].tolist() for m in ("POS", "NEG")}
    npc = {"POS": npc_matrix(pos, order["POS"]), "NEG": npc_matrix(neg, order["NEG"])}

    # sanity: NPC row totals equal the CANOPUS-class counts
    for m, d in [("POS", pos), ("NEG", neg)]:
        assert npc[m].sum().sum() == int(as_bool(d["has_canopus_class"]).sum()), \
            f"{m}: NPC matrix total != CANOPUS count"
    print("GATE PASS - NPC matrix totals == CANOPUS-class counts (POS & NEG)")

    OUTDIR.mkdir(parents=True)
    _write_workbook(OUT_XLSX, pos, neg, pp, npc)

    for m in ("POS", "NEG"):
        pp[m].to_csv(OUTDIR / f"S9_coverage_by_phylum_{m}.csv", index=False)
        npc[m].to_csv(OUTDIR / f"S9_npc_by_phylum_{m}.csv")
    summary = {
        "table": "S9", "level": "per strict phylum, POS + NEG",
        "source": "annotation/annotation_evidence_{pos,neg}.csv (integrated LISC "
                  "jobs 5800445/5802297 + historical atlas SIRIUS cache)",
        "n_biomarkers": {"POS": len(pos), "NEG": len(neg)},
        "overall_rates": {
            m: {label: {"n": int(as_bool(d[col]).sum()),
                        "pct": round(100 * as_bool(d[col]).sum() / len(d), 1)}
                for label, col in METRICS}
            for m, d in [("POS", pos), ("NEG", neg)]},
        "consistent_with": ["Figure 2", "Table S5", "Supplementary Figure 7"],
        "producer": "paper2_repro/scripts/build_table_s9_strict16.py",
    }
    (OUTDIR / "RUN_SUMMARY.json").write_text(json.dumps(summary, indent=2) + "\n",
                                             encoding="utf-8")
    print(f"\nWrote {OUT_XLSX}")
    return 0


def _title(ws, text, ncol):
    c = ws.cell(row=1, column=1, value=text); c.font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)


def _coverage_by_phylum_sheet(ws, title, df):
    _title(ws, title, 3 + 2 * len(METRICS))
    hdr = ["Phylum", "Kingdom", "n biomarkers"]
    for label, _ in METRICS:
        hdr += [f"{label} (n)", f"{label} %"]
    for j, name in enumerate(hdr, start=1):
        c = ws.cell(row=2, column=j, value=name); c.font = H; c.alignment = WRAP
    r0 = 3
    for i, (_, row) in enumerate(df.iterrows()):
        r = r0 + i
        ws.cell(row=r, column=1, value=row["phylum"]).font = BODY_B
        ws.cell(row=r, column=2, value=row["kingdom"]).font = BODY
        ws.cell(row=r, column=3, value=int(row["n"])).font = BODY
        ncol_letter = get_column_letter(3)
        for k, (_, col) in enumerate(METRICS):
            ccol = 4 + 2 * k          # count column
            pcol = 5 + 2 * k          # percent column
            ws.cell(row=r, column=ccol, value=int(row[col])).font = BODY
            pc = ws.cell(row=r, column=pcol,
                         value=f"={get_column_letter(ccol)}{r}/{ncol_letter}{r}")
            pc.font = BODY; pc.number_format = "0.0%"
    # total row (live SUM + rate formulas)
    rt = r0 + len(df)
    ws.cell(row=rt, column=1, value="All phyla").font = BODY_B
    ws.cell(row=rt, column=3,
            value=f"=SUM(C{r0}:C{rt-1})").font = BODY_B
    for k, (_, col) in enumerate(METRICS):
        ccol = 4 + 2 * k; pcol = 5 + 2 * k
        cl = get_column_letter(ccol)
        tc = ws.cell(row=rt, column=ccol, value=f"=SUM({cl}{r0}:{cl}{rt-1})")
        tc.font = BODY_B
        pc = ws.cell(row=rt, column=pcol, value=f"={cl}{rt}/C{rt}")
        pc.font = BODY_B; pc.number_format = "0.0%"
    widths = [16, 14, 12] + [11, 8] * len(METRICS)
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _npc_sheet(ws, title, mat, pp_df):
    ncol = 2 + len(NPC_ORDER) + 1
    _title(ws, title, ncol)
    hdr = ["Phylum", "Kingdom"] + NPC_ORDER + ["CANOPUS classified (n)"]
    for j, name in enumerate(hdr, start=1):
        c = ws.cell(row=2, column=j, value=name); c.font = H; c.alignment = WRAP
    king = dict(zip(pp_df["phylum"], pp_df["kingdom"]))
    r0 = 3
    for i, phy in enumerate(mat.index):
        r = r0 + i
        ws.cell(row=r, column=1, value=phy).font = BODY_B
        ws.cell(row=r, column=2, value=king.get(phy, "")).font = BODY
        for j, path in enumerate(NPC_ORDER, start=3):
            ws.cell(row=r, column=j, value=int(mat.loc[phy, path])).font = BODY
        first = get_column_letter(3); last = get_column_letter(2 + len(NPC_ORDER))
        tc = ws.cell(row=r, column=ncol, value=f"=SUM({first}{r}:{last}{r})")
        tc.font = BODY
    rt = r0 + len(mat)
    ws.cell(row=rt, column=1, value="All phyla").font = BODY_B
    for j in range(3, ncol + 1):
        cl = get_column_letter(j)
        ws.cell(row=rt, column=j, value=f"=SUM({cl}{r0}:{cl}{rt-1})").font = BODY_B
    widths = [16, 14] + [13] * len(NPC_ORDER) + [16]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _write_workbook(path, pos, neg, pp, npc):
    wb = openpyxl.Workbook()
    notes = wb.active; notes.title = "Notes"
    N = len(pos); M = len(neg)
    lines = [
        ("Supplementary Table S9. Computational annotation coverage by strict phylum (SIRIUS / CANOPUS / CSI:FingerID / MSNovelist)", TITLE),
        (f"Per-phylum coverage of the computational annotation pipeline over the strict biomarker set ({N} positive-mode and {M} negative-mode biomarkers, 16 phyla each). SIRIUS predicts a molecular formula; CANOPUS assigns an NPClassifier / ClassyFire compound class; CSI:FingerID proposes a database structure; MSNovelist proposes a de novo structure; DreaMS provides spectral-neighbour evidence.", SUB),
        ("Denominator is all strict biomarkers in each phylum. Rates are live percentage formulas (n / n biomarkers). 'Usable MS2' is the fraction with a usable MS2 spectrum; SIRIUS additionally requires precursor m/z <= 850 Da.", SUB),
        (f"Coverage note: the SIRIUS/CANOPUS annotation is now COMPLETE for the SIRIUS-eligible strict biomarkers. An earlier submission gap - only ~52% of positive-mode and ~69% of negative-mode biomarkers had been submitted to SIRIUS, leaving Bacteria and Fungi near 3% and 10% - was closed by a dedicated gap-fill run of the 2,209 positive-mode and 80 negative-mode eligible-unsubmitted features (usable MS2 and m/z <= 850; LISC jobs 5862224 for formula/CANOPUS/de-novo and 5869613 for CSI:FingerID structures over a BIO+PubChem database union, 2026-08-13). Positive-mode SIRIUS-formula coverage consequently rose from 5,865 to {int(pos['has_sirius_formula'].sum()):,} and CANOPUS from 5,325 to {int(pos['has_canopus_class'].sum()):,}. Remaining non-coverage reflects genuine ineligibility (no usable MS2, or precursor m/z > 850 Da, which is SIRIUS's hard limit), not a submission shortfall; the per-phylum rates below are complete.", SUB),
        ("Sheets: 'Coverage_overall' (pipeline-step totals for both modes); 'Coverage_by_phylum_POS' / '_NEG' (per-phylum counts and rates); 'NPC_pathway_by_phylum_POS' / '_NEG' (NPClassifier pathway x phylum matrix over CANOPUS-classified features).", SUB),
        ("Kingdom labels follow the locked ecological_group vocabulary (Viridiplantae, Protists). Values are drawn from the integrated strict annotation evidence, so this table is consistent with Figure 2, Table S5 and Supplementary Figure 7 by construction.", SUB),
        (f"Overall positive mode: SIRIUS formula {int(pos['has_sirius_formula'].sum())} ({100*pos['has_sirius_formula'].sum()/N:.1f}%), CANOPUS {int(pos['has_canopus_class'].sum())} ({100*pos['has_canopus_class'].sum()/N:.1f}%). Overall negative mode: SIRIUS {int(neg['has_sirius_formula'].sum())} ({100*neg['has_sirius_formula'].sum()/M:.1f}%), CANOPUS {int(neg['has_canopus_class'].sum())} ({100*neg['has_canopus_class'].sum()/M:.1f}%).", SUB),
        ("Source: annotation/annotation_evidence_{pos,neg}.csv, integrating LISC SIRIUS/CANOPUS jobs 5800445 (POS) and 5802297 (NEG) plus the 2026-08-13 gap-fill jobs 5862224 (formula/CANOPUS/de-novo) and 5869613 (CSI:FingerID structures) with the historical full-atlas SIRIUS cache via build_strict_annotation_evidence.py. Producer: paper2_repro/scripts/build_table_s9_strict16.py.", SUB_I),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=text); c.font = font; c.alignment = WRAP
        notes.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    notes.column_dimensions["A"].width = 118

    # Coverage_overall
    ws = wb.create_sheet("Coverage_overall")
    _title(ws, "Supplementary Table S9a. Annotation coverage overall (positive and negative mode)", 5)
    for j, name in enumerate(["Pipeline step", "POS features", "POS rate", "NEG features", "NEG rate"], start=1):
        c = ws.cell(row=2, column=j, value=name); c.font = H; c.alignment = WRAP
    ws.cell(row=3, column=1, value=f"Strict biomarkers (input)").font = BODY_B
    ws.cell(row=3, column=2, value=N).font = BODY_B
    ws.cell(row=3, column=4, value=M).font = BODY_B
    for i, (label, col) in enumerate(METRICS):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = BODY
        ws.cell(row=r, column=2, value=int(pos[col].sum())).font = BODY
        pc = ws.cell(row=r, column=3, value=f"=B{r}/$B$3"); pc.font = BODY; pc.number_format = "0.0%"
        ws.cell(row=r, column=4, value=int(neg[col].sum())).font = BODY
        nc = ws.cell(row=r, column=5, value=f"=D{r}/$D$3"); nc.font = BODY; nc.number_format = "0.0%"
    for j, w in enumerate([30, 14, 10, 14, 10], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    _coverage_by_phylum_sheet(wb.create_sheet("Coverage_by_phylum_POS"),
                              "Supplementary Table S9b. Positive-mode annotation coverage by strict phylum", pp["POS"])
    _coverage_by_phylum_sheet(wb.create_sheet("Coverage_by_phylum_NEG"),
                              "Supplementary Table S9c. Negative-mode annotation coverage by strict phylum", pp["NEG"])
    _npc_sheet(wb.create_sheet("NPC_pathway_by_phylum_POS"),
               "Supplementary Table S9d. Positive-mode NPClassifier pathway by strict phylum (CANOPUS)", npc["POS"], pp["POS"])
    _npc_sheet(wb.create_sheet("NPC_pathway_by_phylum_NEG"),
               "Supplementary Table S9e. Negative-mode NPClassifier pathway by strict phylum (CANOPUS)", npc["NEG"], pp["NEG"])
    wb.save(path)


if __name__ == "__main__":
    sys.exit(main())
