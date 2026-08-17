#!/usr/bin/env python3
"""Supplementary Table S5 - per-phylum biomarker counts by polarity and tier,
built from the strict release biomarker set.

Source: the strict annotation-evidence tables
`annotation/annotation_evidence_pos.csv` (11,371 features) and
`annotation_evidence_neg.csv` (5,697 features). These are the biomarker features
discovered on the locked release's 16 analysis phyla (discovery_method =
indval_consensus + composite), the SAME set Figure 2 and Supplementary Figure 7
are built from - their per-(phylum, tier) counts reproduce
`annotation/tier_counts.csv` exactly.

This is NOT a relabel of the submitted table. The submitted S5 counted the
pre-correction atlas (10,454 POS / 5,741 NEG across 21 / 23 legacy phyla); this
counts a freshly discovered strict biomarker set on 16 units. Relabelling the
old atlas was shown to be invalid - two of three merged units would have been
overstated ~3x (see table_s5_strict16_.../STATUS_BLOCKED.md). The strict
Streptophyta count is 5,168 (POS) versus a legacy land-plant sum of 1,554,
which is exactly why a fresh discovery, not a sum, is required.

Kingdom labels are normalised to the locked policy's ecological_group
(Viridiplantae, Protists) because the two source files disagree - the POS file
already uses the policy names, the NEG file still carries Plantae / Protozoa.

Tier scheme is the annotation release's: Gold (molecular species), Silver
(sum composition / partial), Bronze (lipid class only), Unidentified. Silver is
assigned only where the release supports it (the two archaeal phyla, positive
mode); negative mode has no Silver.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
RELEASE_ROOT = PROJECT_ROOT / "outputs" / "analysis" / "ncbi-phylum-2026-08-04-v1"
ANN = RELEASE_ROOT / "annotation"
TAX = RELEASE_ROOT / "taxonomy"
POLICY = PROJECT_ROOT / "paper2_repro" / "config" / "taxonomy_policy.json"
SUBMITTED = PROJECT_ROOT / "outputs" / "tables" / "Table_S5_biomarker_counts_by_phylum.xlsx"
OUT_DIR = RELEASE_ROOT / "tables" / "table_s5_strict16_2026-08-12_v1"
OUT_XLSX = OUT_DIR / "Table_S5_biomarker_counts_by_phylum_strict16.xlsx"

TIERS = ["Gold", "Silver", "Bronze", "Unidentified"]
FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
TOTAL_FILL = PatternFill("solid", fgColor="EAEAEA")


def counts_for(mode: str, evidence_csv: Path, groups: dict) -> pd.DataFrame:
    d = pd.read_csv(evidence_csv, low_memory=False)
    ct = (pd.crosstab(d["phylum"], d["annotation_tier"])
          .reindex(columns=TIERS).fillna(0).astype(int))
    ct.index.name = "Phylum"
    ct = ct.reset_index()
    ct.insert(1, "Kingdom", ct["Phylum"].map(groups))
    ct["Total"] = ct[TIERS].sum(axis=1)
    ct["Gold+Silver"] = ct["Gold"] + ct["Silver"]
    ct["Annotated"] = ct["Gold"] + ct["Silver"] + ct["Bronze"]
    ct["Annotated %"] = (100 * ct["Annotated"] / ct["Total"]).round(1)
    ct = ct.sort_values(["Kingdom", "Phylum"]).reset_index(drop=True)
    return ct


def main() -> int:
    if OUT_DIR.exists() and OUT_XLSX.exists():
        sys.exit(f"Refusing to overwrite {OUT_XLSX} - delete it first.")

    policy = json.loads(POLICY.read_text(encoding="utf-8"))
    assert policy["status"] == "locked"
    groups = policy["ecological_group"]
    summary = json.loads((TAX / "taxonomy_summary.json").read_text())

    pos = counts_for("POS", ANN / "annotation_evidence_pos.csv", groups)
    neg = counts_for("NEG", ANN / "annotation_evidence_neg.csv", groups)

    # gate: totals must equal the strict feature counts used by the figures
    rel = pd.read_csv(ANN / "annotation_release_summary.csv").set_index("mode")
    assert int(pos["Total"].sum()) == int(rel.loc["POS", "strict_features"]), \
        (pos["Total"].sum(), rel.loc["POS", "strict_features"])
    assert int(neg["Total"].sum()) == int(rel.loc["NEG", "strict_features"])
    tc = pd.read_csv(ANN / "tier_counts.csv")
    for mode, frame in (("POS", pos), ("NEG", neg)):
        long = (frame.melt(id_vars=["Phylum"], value_vars=TIERS,
                           var_name="annotation_tier", value_name="n")
                .query("n > 0").rename(columns={"Phylum": "phylum"})
                .set_index(["phylum", "annotation_tier"])["n"])
        ref = (tc[tc["mode"] == mode].set_index(["phylum", "annotation_tier"])
               ["n_features"])
        j = long.to_frame("mine").join(ref.rename("ref"), how="outer").fillna(0)
        assert (j["mine"] == j["ref"]).all(), f"{mode} disagrees with tier_counts.csv"
    print("GATE PASS - both polarities tie to tier_counts.csv and the release "
          "feature totals")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    cols = ["Phylum", "Kingdom"] + TIERS + ["Total", "Gold+Silver",
                                            "Annotated", "Annotated %"]
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as xw:
        pos[cols].to_excel(xw, sheet_name="Positive_mode", index=False, startrow=3)
        neg[cols].to_excel(xw, sheet_name="Negative_mode", index=False, startrow=3)

    wb = load_workbook(OUT_XLSX)

    # per-sheet: header styling, total row with live formulas
    for sheet, frame in (("Positive_mode", pos), ("Negative_mode", neg)):
        ws = wb[sheet]
        h = {ws.cell(row=4, column=j).value: j for j in range(1, ws.max_column + 1)}
        last = ws.max_row
        trow = last + 1
        ws.cell(row=trow, column=h["Phylum"], value="TOTAL").font = Font(
            name=FONT, size=10, bold=True)
        for col in TIERS + ["Total", "Gold+Silver", "Annotated"]:
            L = get_column_letter(h[col])
            c = ws.cell(row=trow, column=h[col], value=f"=SUM({L}5:{L}{last})")
            c.font = Font(name=FONT, size=10, bold=True)
        # annotated % of the total, as a formula
        La, Lt = get_column_letter(h["Annotated"]), get_column_letter(h["Total"])
        pc = ws.cell(row=trow, column=h["Annotated %"],
                     value=f"=ROUND(100*{La}{trow}/{Lt}{trow},1)")
        pc.font = Font(name=FONT, size=10, bold=True)
        for j in range(1, ws.max_column + 1):
            ws.cell(row=trow, column=j).fill = TOTAL_FILL

    titles = {
        "Positive_mode": (
            f"Supplementary Table S5a. Positive-mode biomarker counts "
            f"({int(pos['Total'].sum()):,} biomarkers, "
            f"{summary['n_analysis_phyla']} analysis phyla)",
            "Strict release biomarker set (indval_consensus + composite "
            "discovery), by annotation tier."),
        "Negative_mode": (
            f"Supplementary Table S5b. Negative-mode biomarker counts "
            f"({int(neg['Total'].sum()):,} biomarkers, "
            f"{summary['n_analysis_phyla']} analysis phyla)",
            "Strict release biomarker set. Negative mode has no Silver tier."),
    }
    for sheet, (t1, t2) in titles.items():
        ws = wb[sheet]
        ws.cell(row=1, column=1, value=t1).font = Font(name=FONT, size=11, bold=True)
        ws.cell(row=2, column=1, value=t2).font = Font(name=FONT, size=10, italic=True)
        for j in range(1, ws.max_column + 1):
            hc = ws.cell(row=4, column=j)
            hc.font = Font(name=FONT, size=10, bold=True)
            hc.fill = HDR_FILL
            hc.alignment = Alignment(wrap_text=True, vertical="bottom",
                                     horizontal="center" if j > 2 else "left")
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row - 1,
                                max_col=ws.max_column):
            for c in row:
                c.font = Font(name=FONT, size=10)
        for j in range(1, ws.max_column + 1):
            L = get_column_letter(j)
            longest = max((len(str(ws.cell(row=i, column=j).value or ""))
                           for i in range(4, ws.max_row + 1)), default=10)
            ws.column_dimensions[L].width = min(max(longest + 2, 8), 20)
        ws.freeze_panes = "A5"

    notes = wb.create_sheet("Notes", 0)
    sub_pos, sub_neg = 10454, 5741
    lines = [
        ("Supplementary Table S5. Per-phylum biomarker counts by polarity and "
         "annotation tier (strict release)", True),
        ("", False),
        (f"Positive mode: {int(pos['Total'].sum()):,} biomarkers across "
         f"{summary['n_analysis_phyla']} analysis phyla. Negative mode: "
         f"{int(neg['Total'].sum()):,} biomarkers across "
         f"{summary['n_analysis_phyla']} phyla.", False),
        ("", False),
        ("Annotation tiers: Gold = molecular-species annotation; Silver = sum "
         "composition or partial structure; Bronze = lipid-class level; "
         "Unidentified = no structural annotation. Silver appears only where the "
         "release supports it (the two archaeal phyla, positive mode); negative "
         "mode has no Silver tier.", False),
        ("", False),
        ("Source: the strict biomarker set in annotation_evidence_pos.csv and "
         "annotation_evidence_neg.csv - the same features Figure 2 and "
         "Supplementary Figure 7 are built from. Per-phylum, per-tier counts "
         "reproduce annotation/tier_counts.csv exactly, and the totals match the "
         "release feature counts (11,371 POS, 5,697 NEG).", False),
        ("", False),
        ("This is a fresh strict discovery, NOT a relabel of the submitted "
         f"table (which counted {sub_pos:,} POS / {sub_neg:,} NEG over the "
         "pre-correction 21 / 23 phyla). Relabelling was shown to be invalid: "
         "two of three merged units would have been overstated roughly "
         "threefold, because a biomarker is defined by contrast against a "
         "background and merging changes that background. The strict "
         f"Streptophyta count ({int(pos.loc[pos.Phylum=='Streptophyta','Total'].iloc[0]):,} "
         "POS) versus a legacy land-plant sum of 1,554 shows why a fresh "
         "discovery, not a sum, is required.", False),
        ("", False),
        ("Kingdom labels use the locked policy's ecological_group naming "
         "(Viridiplantae, Protists). The two source files disagreed - the "
         "positive file already used these, the negative file still carried "
         "Plantae / Protozoa - and are reconciled here.", False),
        ("", False),
        (f"Taxonomy release: {summary['taxonomy_release']} (locked). "
         "Producer: paper2_repro/scripts/build_table_s5_strict16.py", False),
    ]
    for i, (txt, bold) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=txt)
        c.font = Font(name=FONT, size=11, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    notes.column_dimensions["A"].width = 114

    wb.save(OUT_XLSX)

    pos[cols].to_csv(OUT_DIR / "s5_positive_mode_strict16.csv", index=False)
    neg[cols].to_csv(OUT_DIR / "s5_negative_mode_strict16.csv", index=False)
    (OUT_DIR / "RUN_SUMMARY.json").write_text(json.dumps({
        "taxonomy_release": summary["taxonomy_release"],
        "status": "built from the strict biomarker set (annotation_evidence); "
                  "ties to tier_counts.csv and the figures; NOT a relabel",
        "pos_total": int(pos["Total"].sum()),
        "neg_total": int(neg["Total"].sum()),
        "pos_tiers": {t: int(pos[t].sum()) for t in TIERS},
        "neg_tiers": {t: int(neg[t].sum()) for t in TIERS},
        "submitted_totals": {"POS": sub_pos, "NEG": sub_neg},
        "gate": "PASS - both polarities reproduce tier_counts.csv exactly",
    }, indent=2) + "\n", encoding="utf-8")

    print(f"\nPOS ({int(pos['Total'].sum()):,}):")
    print(pos[["Phylum", "Kingdom"] + TIERS + ["Total"]].to_string(index=False))
    print(f"\nNEG ({int(neg['Total'].sum()):,}):")
    print(neg[["Phylum", "Kingdom"] + TIERS + ["Total"]].to_string(index=False))
    print(f"\nWrote {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
