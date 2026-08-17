#!/usr/bin/env python3
"""Table S15 (five-rank taxonomic framework) on the strict phyla.

Ordinal (cladistic) framework Domain > Supergroup > Organism group (kingdom) >
Clade > Phylum. This matrix is the phylogenetic axis of Supplementary
Figures 1–2 and Table S4. It is NOT the Figure 3b x-axis (that is SSU
rRNA patristic distance; methods/09_phylogeny/). Pairwise
distance = number of ranks from a phylum up to the most recent common ancestor,
i.e. 5 - (matching leading ranks in [Domain, Supergroup, Organism group, Clade]);
0 same phylum ... 5 different domain.

Scope: all 19 collection phyla; the 'In 16-phylum analysis set' column marks the
16 with n>=2 samples in both polarities. Re-derived on strict units (Amoebozoa ->
Discosea/Evosea as distinct clades; land plants + Charophyta -> Streptophyta;
Euryarchaeota -> Methanobacteriota; Crenarchaeota -> Thermoproteota). Streptophyta
clade PENDING coauthor confirmation (spans charophyte algae + land plants).
Opisthokonta retained as the eukaryotic supergroup uniting Fungi and Animalia.

GATE: the rank-to-MRCA distances recomputed from this framework must reproduce the
strict 16-phylum matrix `phylo_dist_16phyla.csv` EXACTLY (max |delta| = 0).
"""

from __future__ import annotations

import json
import sys
from itertools import combinations
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
RE = ROOT / "outputs/analysis/ncbi-phylum-2026-08-04-v1"
PHYLO = RE / "suppfig1_full_strict16_2026-08-11_v1/panel_d_data/phylo_dist_16phyla.csv"
OUTDIR = RE / "tables/table_s15_strict16_2026-08-12_v1"
OUT_XLSX = OUTDIR / "Table_S15_phylogeny_cladogram_strict16.xlsx"

# (Domain, Supergroup, Organism group, Clade, Phylum, in_16_analysis_set)
FRAMEWORK = [
    ("Bacteria", "Terrabacteria", "Bacteria", "Actinobacteria", "Actinomycetota", True),
    ("Bacteria", "Terrabacteria", "Bacteria", "Firmicutes", "Bacillota", True),
    ("Bacteria", "Terrabacteria", "Bacteria", "Cyanobacteria", "Cyanobacteriota", False),
    ("Bacteria", "Gracilicutes", "Bacteria", "Proteobacteria", "Pseudomonadota", True),
    ("Archaea", "Euryarchaeota", "Archaea", "Methanobacteria", "Methanobacteriota", True),
    ("Archaea", "TACK", "Archaea", "Thermoproteota", "Thermoproteota", True),
    ("Eukaryota", "Opisthokonta", "Fungi", "Dikarya", "Ascomycota", True),
    ("Eukaryota", "Opisthokonta", "Fungi", "Dikarya", "Basidiomycota", True),
    ("Eukaryota", "Opisthokonta", "Fungi", "Basal Fungi", "Mucoromycota", True),
    ("Eukaryota", "Opisthokonta", "Fungi", "Basal Fungi", "Mortierellomycota", False),
    ("Eukaryota", "Opisthokonta", "Animalia", "Ecdysozoa", "Arthropoda", True),
    ("Eukaryota", "Opisthokonta", "Animalia", "Ecdysozoa", "Nematoda", True),
    ("Eukaryota", "Opisthokonta", "Animalia", "Lophotrochozoa", "Mollusca", True),
    ("Eukaryota", "Archaeplastida", "Viridiplantae", "Chlorophyta", "Chlorophyta", True),
    ("Eukaryota", "Archaeplastida", "Viridiplantae", "Streptophyta", "Streptophyta", True),
    ("Eukaryota", "Amoebozoa", "Protists", "Discosea", "Discosea", True),
    ("Eukaryota", "Amoebozoa", "Protists", "Evosea", "Evosea", True),
    ("Eukaryota", "SAR", "Protists", "Rhizaria", "Cercozoa", False),
    ("Eukaryota", "Excavata", "Protists", "Discoba", "Heterolobosea", True),
]
COLS = ["Domain", "Supergroup", "Organism group", "Clade", "Phylum", "In 16-phylum analysis set"]

TITLE = Font(name="Arial", size=13, bold=True)
SUB = Font(name="Arial", size=10)
SUB_I = Font(name="Arial", size=9, italic=True)
H = Font(name="Arial", size=10, bold=True)
BODY = Font(name="Arial", size=10)
BODY_B = Font(name="Arial", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def rank_distance(a: tuple, b: tuple) -> int:
    """5 - number of matching leading ranks in [Domain, Supergroup, Org, Clade]."""
    if a[4] == b[4]:
        return 0
    match = 0
    for i in range(4):  # Domain, Supergroup, Organism group, Clade
        if a[i] == b[i]:
            match += 1
        else:
            break
    return 5 - match


def main() -> int:
    if OUTDIR.exists():
        sys.exit(f"Refusing to overwrite {OUTDIR} - delete or rename it first.")

    rows = [dict(zip(["domain", "supergroup", "org", "clade", "phylum", "ana"], r))
            for r in FRAMEWORK]
    by_phy = {r["phylum"]: tuple(FRAMEWORK[i]) for i, r in enumerate(rows)}
    analysis16 = [r["phylum"] for r in rows if r["ana"]]
    assert len(FRAMEWORK) == 19 and len(analysis16) == 16, \
        f"expected 19 collection / 16 analysis, got {len(FRAMEWORK)}/{len(analysis16)}"

    # ---- GATE: reproduce phylo_dist_16phyla.csv exactly ----------------------
    phylo = pd.read_csv(PHYLO, index_col=0)
    assert set(phylo.index) == set(analysis16), \
        f"matrix phyla != analysis set: {set(phylo.index) ^ set(analysis16)}"
    maxd = 0
    bad = []
    for a, b in combinations(analysis16, 2):
        d_mine = rank_distance(by_phy[a], by_phy[b])
        d_ref = int(phylo.loc[a, b])
        if d_mine != d_ref:
            bad.append((a, b, d_mine, d_ref))
        maxd = max(maxd, abs(d_mine - d_ref))
    assert not bad, f"framework distances disagree with phylo_dist_16phyla.csv: {bad[:8]}"
    print(f"GATE PASS - rank-to-MRCA distances reproduce phylo_dist_16phyla.csv "
          f"exactly for all {len(analysis16)} analysis phyla (max |delta| {maxd})")

    OUTDIR.mkdir(parents=True)
    wb = openpyxl.Workbook()
    notes = wb.active; notes.title = "Notes"
    lines = [
        ("Supplementary Table S15. Five-rank taxonomic framework (strict phyla)", TITLE),
        ("Ordinal (cladistic) framework underlying the phylum phylogenetic distance matrix used in the lipidome-phylogeny Mantel and partial-Mantel tests (Fig. 3, Fig. 4). Ranks, broad to narrow: Domain > Supergroup > Organism group (kingdom) > Clade > Phylum.", SUB),
        ("Pairwise distance = number of ranks from a phylum up to the most recent common ancestor: 0 = same phylum; 1 = same clade; 2 = same organism group; 3 = same supergroup; 4 = same domain; 5 = different domain. Equivalently, 5 minus the number of matching leading ranks. The distances recomputed from this framework reproduce the strict 16-phylum matrix (phylo_dist_16phyla.csv) exactly.", SUB),
        ("Scope: all 19 collection phyla. The 'In 16-phylum analysis set' column marks the 16 with n>=2 samples in both polarities used for phylum inference; the other three (Cyanobacteriota, Mortierellomycota, Cercozoa) are below threshold and retained for completeness only.", SUB),
        ("Re-derived on strict units: Amoebozoa split into Discosea and Evosea (distinct clades within the Amoebozoa supergroup, distance 2); the legacy land-plant labels and Charophyta merged into Streptophyta; Euryarchaeota into Methanobacteriota; Crenarchaeota into Thermoproteota. Organism-group labels follow the locked ecological_group vocabulary (Viridiplantae, Protists).", SUB),
        ("PENDING coauthor confirmation: the Streptophyta clade placement. The strict Streptophyta phylum spans charophyte algae (Charophyta) and land plants (Embryophyta); it is assigned the superclade 'Streptophyta' here, distinct from Chlorophyta (distance 2). Opisthokonta is retained as the eukaryotic supergroup uniting Fungi and Animalia.", SUB),
        ("Basis: eukaryotic supergroups follow Adl et al. (2019, J. Eukaryot. Microbiol. 66:4-119); rank/name assignments follow NCBI Taxonomy (Schoch et al. 2020). Source: paper2_repro/scripts/build_table_s15_strict16.py, gated against suppfig1_full_strict16_2026-08-11_v1/panel_d_data/phylo_dist_16phyla.csv.", SUB_I),
    ]
    for i, (t, f) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=t); c.font = f; c.alignment = WRAP
        notes.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    notes.column_dimensions["A"].width = 118

    ws = wb.create_sheet("Cladogram")
    c = ws.cell(row=1, column=1, value="Supplementary Table S15. Five-rank taxonomic framework (strict phyla)")
    c.font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    for j, name in enumerate(COLS, start=1):
        hc = ws.cell(row=2, column=j, value=name); hc.font = H; hc.alignment = WRAP
    for i, r in enumerate(FRAMEWORK, start=3):
        vals = list(r[:5]) + ["yes" if r[5] else "no (below threshold)"]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = BODY_B if j == 5 else BODY
            if j == 4 and r[4] == "Streptophyta":
                cell.value = "Streptophyta *"   # pending-confirmation marker
    foot = ws.cell(row=len(FRAMEWORK) + 3, column=1,
                   value="* Streptophyta clade placement pending coauthor confirmation (spans Charophyta + Embryophyta).")
    foot.font = SUB_I
    ws.merge_cells(start_row=len(FRAMEWORK) + 3, start_column=1, end_row=len(FRAMEWORK) + 3, end_column=len(COLS))
    for j, w in enumerate([12, 15, 16, 16, 18, 22], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    wb.save(OUT_XLSX)

    pd.DataFrame(FRAMEWORK, columns=["Domain", "Supergroup", "Organism_group", "Clade",
                                     "Phylum", "in_16_analysis"]).to_csv(
        OUTDIR / "S15_framework.csv", index=False)
    (OUTDIR / "RUN_SUMMARY.json").write_text(json.dumps({
        "table": "S15", "title": "Five-rank taxonomic framework",
        "n_collection_phyla": 19, "n_analysis_phyla": 16,
        "distance_gate": "PASS - reproduces phylo_dist_16phyla.csv exactly (max |delta| 0)",
        "decisions": {"streptophyta_clade": "Streptophyta (PENDING coauthor confirmation)",
                      "opisthokonta": "retained as eukaryotic supergroup (Fungi+Animalia)",
                      "scope": "19 collection phyla, 16-analysis flag"},
        "below_threshold": ["Cyanobacteriota", "Mortierellomycota", "Cercozoa"],
    }, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
