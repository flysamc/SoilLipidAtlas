#!/usr/bin/env python3
"""Table S10 (ClimGrass microbial benchmark studies) on the locked release naming.

S10 is a LITERATURE table: four independent published studies from the ClimGrass
platform (Deltedesco 2020, Metze 2023, Canarini 2024, Seneca 2020) used as
qualitative site-level plausibility benchmarks. No strict-taxonomy recomputation
applies.

Two rules:
  * Taxon names in the 'Main finding' column (Proteobacteria, Verrucomicrobia,
    Actinobacteria / Actinobacteriota, Acidobacteria, Ascomycota) QUOTE the cited
    studies and are preserved verbatim -- remapping them to NCBI phyla would
    misrepresent the sources.
  * The one word that is OUR own label -- "Protozoa" in the Notes ("no benchmark
    for the Animalia or Protozoa fractions") -- is conformed to the locked
    ecological_group vocabulary "Protists" (as Fig 5 v2 and Table S8 use).

Edit-in-place on the frozen submitted workbook so all formatting/refs are kept;
gates = reproduce-first on the Protozoa anchor, cited terms preserved, exactly one
pre-existing cell changed.
"""

from __future__ import annotations

import csv
import json
import re
import sys
from copy import copy
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "outputs" / "tables" / "Table_S10_climgrass_benchmark_studies.xlsx"
RELEASE = ROOT / "outputs" / "analysis" / "ncbi-phylum-2026-08-04-v1"
OUTDIR = RELEASE / "tables" / "table_s10_strict16_2026-08-12_v1"
OUT_XLSX = OUTDIR / "Table_S10_climgrass_benchmark_studies_strict16.xlsx"

CITED_TERMS = ["Proteobacteria", "Verrucomicrobia", "Actinobacteria",
               "Acidobacteria", "Actinobacteriota", "Ascomycota"]
ANCHOR = ("Notes", "A4")  # the "...Animalia or Protozoa fractions" note


def snapshot(wb) -> dict:
    return {(ws.title, c.coordinate): c.value
            for ws in wb.worksheets for row in ws.iter_rows()
            for c in row if c.value is not None}


def main() -> int:
    if OUTDIR.exists():
        sys.exit(f"Refusing to overwrite {OUTDIR} - delete or rename it first.")

    wb = openpyxl.load_workbook(SRC)

    # ---- reproduce-first gate ------------------------------------------------
    anchor = wb[ANCHOR[0]][ANCHOR[1]]
    assert anchor.value and "Protozoa" in str(anchor.value), \
        f"reproduce-first FAIL: {ANCHOR} does not contain 'Protozoa'"
    protozoa_cells = [(ws.title, c.coordinate) for ws in wb.worksheets
                      for row in ws.iter_rows() for c in row
                      if c.value and "Protozoa" in str(c.value)]
    assert protozoa_cells == [ANCHOR], f"unexpected 'Protozoa' cells: {protozoa_cells}"
    tbl_text = " ".join(str(c.value) for row in wb["Table"].iter_rows()
                        for c in row if c.value)
    for t in CITED_TERMS:
        assert t in tbl_text, f"cited term missing before edit: {t}"
    before = snapshot(wb)
    print(f"Reproduce-first gate PASS - single 'Protozoa' at {ANCHOR}; "
          f"{len(CITED_TERMS)} cited terms present.")

    # ---- apply the one label conformance -------------------------------------
    anchor.value = re.sub(r"Protozoa", "Protists", str(anchor.value))

    # cited terms must be untouched
    tbl_text_after = " ".join(str(c.value) for row in wb["Table"].iter_rows()
                              for c in row if c.value)
    for t in CITED_TERMS:
        assert t in tbl_text_after, f"cited term lost after edit: {t}"

    # ---- provenance note ------------------------------------------------------
    notes = wb["Notes"]
    last = max((c.row for row in notes.iter_rows() for c in row
                if c.value not in (None, "")), default=1)
    prov_row = last + 2
    prov = (
        "Strict-16 update 2026-08-12: 'Protozoa' conformed to the locked "
        "ecological_group label 'Protists' (our composition fraction; consistent "
        "with Fig 5 v2 and Table S8). Taxon names in the 'Main finding' column are "
        "preserved verbatim as reported by the cited studies (Proteobacteria, "
        "Verrucomicrobia, Actinobacteria / Actinobacteriota, Acidobacteria, "
        "Ascomycota) and are NOT remapped to NCBI phyla. This is a literature "
        "benchmark table with no strict-taxonomy recomputation."
    )
    cell = notes.cell(row=prov_row, column=1, value=prov)
    template = notes.cell(row=last, column=1)
    cell.font = copy(template.font)
    cell.alignment = copy(template.alignment)
    notes.merge_cells(start_row=prov_row, start_column=1, end_row=prov_row, end_column=8)

    # ---- verify --------------------------------------------------------------
    after = snapshot(wb)
    changed = {k for k in before if before[k] != after.get(k)}
    dropped = {k for k in before if k not in after}
    added = {k for k in after if k not in before}
    assert dropped == set(), f"cells lost: {dropped}"
    assert changed == {ANCHOR}, f"unexpected changes: {changed}"
    assert added == {("Notes", f"A{prov_row}")}, f"unexpected additions: {added}"
    print(f"Verify PASS - changed exactly {ANCHOR} (Protozoa->Protists); "
          f"provenance at Notes!A{prov_row}; cited terms intact; nothing dropped.")

    # ---- write ---------------------------------------------------------------
    OUTDIR.mkdir(parents=True)
    wb.save(OUT_XLSX)
    with (OUTDIR / "Table_S10_strict16.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        for row in wb["Table"].iter_rows(values_only=True):
            r = list(row)
            while r and r[-1] is None:
                r.pop()
            if any(v is not None for v in r):
                w.writerow(r)
    (OUTDIR / "RUN_SUMMARY.json").write_text(json.dumps({
        "table": "S10",
        "title": "ClimGrass microbial benchmark studies",
        "type": "literature benchmark (no strict-taxonomy recomputation)",
        "relabels_applied": {"Protozoa": "Protists"},
        "cells_changed": [f"{ANCHOR[0]}!{ANCHOR[1]}"],
        "cited_terms_preserved_verbatim": CITED_TERMS,
        "provenance_note_cell": f"Notes!A{prov_row}",
        "gates": {"reproduce_first": "PASS", "exactly_one_change": "PASS",
                  "cited_terms_intact": "PASS"},
    }, indent=2) + "\n", encoding="utf-8")
    print(f"\nWrote {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
