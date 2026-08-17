#!/usr/bin/env python3
"""Table S8 (ClimGrass per-treatment composition) rebuilt on Figure 5 v2.

Built entirely from the Fig 5 v2 outputs on the strict phyla; the deprecated
submitted table is not consulted. Sub-tables:

  S8a  Kingdom x treatment      (mean % +/- SD, 4 treatments, n=3)
  S8b  Phylum x treatment       (16 strict phyla)
  S8c  Treatment effects        (CLR fingerprint-set permutation tests +
                                 qSIP replication family) -- replaces the old
                                 Mann-Whitney-on-fractions, invalid at n=6
  S8d  Per-sample composition   (12 soil samples, fc-weighted, primary)
  S8e  Per-sample composition   (12 soil samples, marker-panel, provenance)

Primary composition = fc-weighted BC with rules (the figure's bars); marker-panel
is the provenance overlay (the figure's diamonds). Kingdom labels follow the
locked ecological_group vocabulary (Viridiplantae / Protists).

Internal gate (within the new analysis, never against the submission): the
fc-weighted phylum fractions must aggregate to the figure's kingdom composition
(r_render/data/kingdom_composition.csv) exactly.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parents[2]
POLICY = ROOT / "paper2_repro" / "config" / "taxonomy_policy.json"
V2 = (ROOT / "outputs" / "analysis" / "ncbi-phylum-2026-08-04-v1" / "climgrass"
      / "figure5_redesign_2026-08-08_v2_archlips")
OUTDIR = (ROOT / "outputs" / "analysis" / "ncbi-phylum-2026-08-04-v1" / "tables"
          / "table_s8_strict16_2026-08-12_v1")
OUT_XLSX = OUTDIR / "Table_S8_climgrass_per_treatment_composition_v2.xlsx"

TREATMENTS = ["Ambient_Control", "Ambient_Drought", "Future_Control", "Future_Drought"]
KING_ORDER = ["Bacteria", "Archaea", "Fungi", "Viridiplantae", "Animalia", "Protists"]

TITLE = Font(name="Arial", size=13, bold=True)
SUB = Font(name="Arial", size=10)
SUB_I = Font(name="Arial", size=9, italic=True)
H = Font(name="Arial", size=10, bold=True)
BODY = Font(name="Arial", size=10)
BODY_B = Font(name="Arial", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
RIGHT = Alignment(horizontal="right")


def load_ecogroup() -> dict:
    pol = json.loads(POLICY.read_text(encoding="utf-8"))
    return dict(pol["ecological_group"])  # phylum -> kingdom (Viridiplantae/Protists)


def main() -> int:
    if OUTDIR.exists():
        sys.exit(f"Refusing to overwrite {OUTDIR} - delete or rename it first.")

    eco = load_ecogroup()

    # ---- load v2 outputs ------------------------------------------------------
    fcw = pd.read_csv(V2 / "composition_fcweighted_by_sample.csv").rename(
        columns={"Unnamed: 0": "column_name"})
    mkr = pd.read_csv(V2 / "composition_winner_by_sample.csv").rename(
        columns={"Unnamed: 0": "column_name"})
    king_fig = pd.read_csv(V2 / "r_render" / "data" / "kingdom_composition.csv")
    fx = pd.read_csv(V2 / "fingerprint_set_effects.csv")
    qsip = pd.read_csv(V2 / "qsip_replication_test.csv")

    phyla = [c for c in fcw.columns if c != "column_name"]
    assert len(phyla) == 16, f"expected 16 strict phyla, got {len(phyla)}"
    for p in phyla:
        assert p in eco, f"phylum {p} missing from ecological_group policy"
    print(f"16 strict phyla: {phyla}")

    # treatment metadata from the figure's kingdom table
    meta = king_fig[["column_name", "treatment", "climate", "drought"]].copy()
    fcw = meta.merge(fcw, on="column_name")
    mkr = meta.merge(mkr, on="column_name")
    assert len(fcw) == 12 and len(mkr) == 12, "expected 12 soil samples"
    tcounts = fcw["treatment"].value_counts().to_dict()
    assert all(tcounts.get(t) == 3 for t in TREATMENTS), f"treatment n!=3: {tcounts}"
    print(f"treatment counts: {tcounts}")

    # ---- kingdom aggregation (phylum -> kingdom) ------------------------------
    def to_kingdom(df: pd.DataFrame) -> pd.DataFrame:
        out = pd.DataFrame({"column_name": df["column_name"],
                            "treatment": df["treatment"],
                            "climate": df["climate"], "drought": df["drought"]})
        for k in KING_ORDER:
            members = [p for p in phyla if eco[p] == k]
            out[k] = df[members].sum(axis=1) if members else 0.0
        return out

    fcw_king = to_kingdom(fcw)
    mkr_king = to_kingdom(mkr)

    # ---- GATE: fc-weighted kingdom agg must match the figure's kingdom table --
    fig = king_fig.set_index("column_name")
    fig_king_names = {"Viridiplantae": "Plantae", "Protists": "Protozoa"}
    maxd = 0.0
    for _, r in fcw_king.iterrows():
        for k in KING_ORDER:
            figcol = fig_king_names.get(k, k)
            maxd = max(maxd, abs(r[k] - fig.loc[r["column_name"], figcol]))
    assert maxd < 1e-6, f"kingdom aggregation does not match figure (maxd {maxd})"
    print(f"GATE PASS - fc-weighted kingdom aggregation matches figure "
          f"(max |delta| {maxd:.1e}); Viridiplantae<-Plantae, Protists<-Protozoa")

    # composition closure gate
    for name, df in [("fcw", fcw), ("mkr", mkr)]:
        s = df[phyla].sum(axis=1)
        assert (s.sub(1.0).abs() < 1e-6).all(), f"{name} rows do not sum to 1"
    print("GATE PASS - all per-sample phylum fractions sum to 1.0")

    # ---- build per-treatment mean/SD tables -----------------------------------
    def by_treatment(df: pd.DataFrame, cats: list[str]) -> pd.DataFrame:
        rows = []
        for cat in cats:
            row = {"Group": cat}
            for t in TREATMENTS:
                sub = df[df["treatment"] == t][cat] * 100.0
                row[f"{t} mean %"] = round(sub.mean(), 2)
                row[f"{t} SD"] = round(sub.std(ddof=1), 2)
            row["Overall mean %"] = round(df[cat].mean() * 100.0, 2)
            rows.append(row)
        return pd.DataFrame(rows)

    s8a = by_treatment(fcw_king, KING_ORDER)
    phyla_by_king = sorted(phyla, key=lambda p: (KING_ORDER.index(eco[p]), p))
    s8b = by_treatment(fcw, phyla_by_king)

    # ---- S8c treatment effects (permutation fingerprint-set tests) ------------
    fx = fx.copy()
    fx["kingdom"] = fx["kingdom"].replace({"Plantae": "Viridiplantae",
                                           "Protozoa": "Protists"})
    fx["direction"] = np.where(fx["set_mean_log2fc"] >= 0, "up", "down")
    fx["p_one_sided"] = np.where(fx["set_mean_log2fc"] >= 0,
                                 fx["p_one_sided_pos"], fx["p_one_sided_neg"])
    s8c = fx[["factor", "unit", "kingdom", "n_set_features", "set_mean_log2fc",
              "direction", "p_perm", "p_one_sided", "q_bh"]].copy()
    s8c = s8c.sort_values(["factor", "unit"]).reset_index(drop=True)

    OUTDIR.mkdir(parents=True)
    _write_workbook(OUT_XLSX, s8a, s8b, s8c, qsip, fcw, fcw_king, mkr, mkr_king,
                    phyla_by_king)

    # ---- CSV mirrors + summary ------------------------------------------------
    s8a.to_csv(OUTDIR / "S8a_kingdom_by_treatment.csv", index=False)
    s8b.to_csv(OUTDIR / "S8b_phylum_by_treatment.csv", index=False)
    s8c.to_csv(OUTDIR / "S8c_treatment_effects.csv", index=False)
    summary = {
        "table": "S8", "rebuilt_for": "Figure 5 v2 (strict phyla)",
        "primary_estimator": "fc_weighted_bc (figure bars)",
        "secondary_estimator": "marker_panel (figure diamonds)",
        "n_phyla": 16, "n_samples": 12, "treatments": TREATMENTS,
        "s8c_test": "CLR fingerprint-set permutation test (400 relabelings) + "
                    "qSIP replication family; replaces Mann-Whitney-on-fractions",
        "qsip_headline": {"Pseudomonadota": {"log2fc": -0.595, "p_one_sided": 0.0025,
                                             "q_bh_family2": 0.005}},
        "kingdom_gate": "fc-weighted phylum aggregation == figure kingdom_composition (max delta <1e-6)",
        "sources": [
            "composition_fcweighted_by_sample.csv", "composition_winner_by_sample.csv",
            "r_render/data/kingdom_composition.csv", "fingerprint_set_effects.csv",
            "qsip_replication_test.csv",
        ],
    }
    (OUTDIR / "RUN_SUMMARY.json").write_text(json.dumps(summary, indent=2) + "\n",
                                             encoding="utf-8")
    print(f"\nWrote {OUT_XLSX}")
    return 0


def _hdr_row(ws, r, names, widths=None):
    for j, name in enumerate(names, start=1):
        c = ws.cell(row=r, column=j, value=name)
        c.font = H; c.alignment = WRAP
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w


def _title(ws, text, ncol):
    c = ws.cell(row=1, column=1, value=text); c.font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)


def _write_treatment_sheet(ws, title, df):
    _title(ws, title, len(df.columns))
    _hdr_row(ws, 2, list(df.columns),
             widths=[16] + [13] * (len(df.columns) - 1))
    for i, (_, row) in enumerate(df.iterrows(), start=3):
        for j, col in enumerate(df.columns, start=1):
            c = ws.cell(row=i, column=j, value=row[col])
            c.font = BODY_B if j == 1 else BODY
            if j > 1:
                c.number_format = "0.00"


def _write_persample_sheet(ws, title, comp, king, phyla_order, king_order):
    cols = (["column_name", "treatment", "climate", "drought"]
            + king_order + phyla_order + ["Sum check"])
    _title(ws, title, len(cols))
    _hdr_row(ws, 2, cols, widths=[26, 16, 9, 11] + [12] * (len(king_order) + len(phyla_order)) + [11])
    kmap = king.set_index("column_name")
    n_meta = 4
    first_king_col = n_meta + 1
    last_phyl_col = n_meta + len(king_order) + len(phyla_order)
    for i, (_, row) in enumerate(comp.iterrows(), start=3):
        vals = [row["column_name"], row["treatment"], row["climate"], row["drought"]]
        for k in king_order:
            vals.append(float(kmap.loc[row["column_name"], k]))  # full precision; format handles display
        for p in phyla_order:
            vals.append(float(row[p]))
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = BODY
            if j > n_meta:
                c.number_format = "0.000000"
        # live sum-check over the phylum columns only (should be 1.0)
        pstart = openpyxl.utils.get_column_letter(n_meta + len(king_order) + 1)
        pend = openpyxl.utils.get_column_letter(last_phyl_col)
        cc = ws.cell(row=i, column=len(vals) + 1,
                     value=f"=SUM({pstart}{i}:{pend}{i})")
        cc.font = BODY; cc.number_format = "0.000000"


def _write_workbook(path, s8a, s8b, s8c, qsip, fcw, fcw_king, mkr, mkr_king, phyla_order):
    wb = openpyxl.Workbook()
    notes = wb.active; notes.title = "Notes"
    lines = [
        ("Supplementary Table S8. ClimGrass per-treatment composition (strict phyla, Figure 5 v2)", TITLE),
        ("Lipid-derived composition for the four ClimGrass treatments (ambient control, ambient + drought, future climate, future + drought; n = 3 each), from the Figure 5 v2 decomposition on the 16 strict analysis phyla.", SUB),
        ("Primary composition = fold-change-weighted Bray-Curtis with hygiene rules (Rule A calibration gate, enriched-weight split, sqrt), the living-community estimate shown as the figure's bars. Marker-panel composition (the provenance-of-signal estimate, shown as the figure's diamonds) is given per sample in sheet S8e.", SUB),
        ("Sheets: 'Kingdom_by_treatment' (S8a) and 'Phylum_by_treatment' (S8b) give per-treatment mean % and SD (n=3). 'Treatment_effects' (S8c) gives drought and climate contrasts from CLR-centred fingerprint-set permutation tests (400 stratified relabelings) with BH q, plus the pre-specified qSIP replication family. 'Per_sample_fcweighted' (S8d) and 'Per_sample_markerpanel' (S8e) give the 12 individual soil samples.", SUB),
        ("Treatment contrasts are exploratory (n = 3 per treatment; n = 6 per side). Composition values are fractions of total decomposed signal; by-treatment sheets express them as percentages. Kingdom labels follow the locked ecological_group vocabulary (Viridiplantae, Protists).", SUB),
        ("Headline drought result (qSIP replication): Pseudomonadota decreases under drought, observed log2FC -0.60, one-sided permutation p = 0.0025, BH q = 0.005 (family of 2). This is a pre-specified directional replication of Metze et al. 2023 qSIP predictions.", SUB),
        ("Source: outputs/analysis/ncbi-phylum-2026-08-04-v1/climgrass/figure5_redesign_2026-08-08_v2_archlips/ (composition_fcweighted_by_sample.csv, composition_winner_by_sample.csv, r_render/data/kingdom_composition.csv, fingerprint_set_effects.csv, qsip_replication_test.csv). Producer: paper2_repro/scripts/build_table_s8_strict16.py.", SUB_I),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=text); c.font = font; c.alignment = WRAP
        notes.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    notes.column_dimensions["A"].width = 115

    _write_treatment_sheet(wb.create_sheet("Kingdom_by_treatment"),
                           "Supplementary Table S8a. Kingdom composition by treatment (mean % +/- SD, n=3)", s8a)
    _write_treatment_sheet(wb.create_sheet("Phylum_by_treatment"),
                           "Supplementary Table S8b. Phylum composition by treatment (mean % +/- SD, n=3)", s8b)

    # S8c treatment effects + qSIP block
    ws = wb.create_sheet("Treatment_effects")
    _title(ws, "Supplementary Table S8c. Treatment effects - CLR fingerprint-set permutation tests", 9)
    sub = ws.cell(row=2, column=1, value="Per-phylum fingerprint-set effects for the drought and climate contrasts (400 stratified CLR relabelings). log2FC = set-mean log2 fold change; p_perm two-sided; p_one_sided in the observed direction; q = Benjamini-Hochberg within factor.")
    sub.font = SUB_I; sub.alignment = WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    cols = ["Factor", "Phylum", "Kingdom", "n set features", "set-mean log2FC",
            "direction", "p_perm (2-sided)", "p one-sided", "q (BH)"]
    _hdr_row(ws, 3, cols, widths=[10, 16, 14, 13, 15, 10, 15, 13, 10])
    r = 4
    for _, row in s8c.iterrows():
        for j, val in enumerate([row["factor"], row["unit"], row["kingdom"],
                                 int(row["n_set_features"]),
                                 round(row["set_mean_log2fc"], 3), row["direction"],
                                 round(row["p_perm"], 4), round(row["p_one_sided"], 4),
                                 round(row["q_bh"], 4)], start=1):
            c = ws.cell(row=r, column=j, value=val); c.font = BODY
            if j in (5, 7, 8, 9):
                c.number_format = "0.000" if j == 5 else "0.0000"
        r += 1
    # qSIP block
    r += 1
    c = ws.cell(row=r, column=1, value="qSIP replication family (pre-specified directional test, Metze et al. 2023)")
    c.font = BODY_B; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9); r += 1
    _hdr_row(ws, r, ["Phylum", "predicted direction", "observed log2FC",
                     "p one-sided", "q (BH, family=2)"]); r += 1
    for _, row in qsip.iterrows():
        for j, val in enumerate([row["unit"], row["predicted_direction"],
                                 round(row["observed_log2fc"], 3),
                                 round(row["p_one_sided"], 4),
                                 round(row["q_bh_family2"], 4)], start=1):
            c = ws.cell(row=r, column=j, value=val); c.font = BODY
            if j in (3, 4, 5):
                c.number_format = "0.000" if j == 3 else "0.0000"
        r += 1

    _write_persample_sheet(wb.create_sheet("Per_sample_fcweighted"),
                           "Supplementary Table S8d. Per-sample composition - fc-weighted (primary), fractions 0-1",
                           fcw, fcw_king, phyla_order, KING_ORDER)
    _write_persample_sheet(wb.create_sheet("Per_sample_markerpanel"),
                           "Supplementary Table S8e. Per-sample composition - marker-panel (provenance overlay), fractions 0-1",
                           mkr, mkr_king, phyla_order, KING_ORDER)
    wb.save(path)


if __name__ == "__main__":
    sys.exit(main())
